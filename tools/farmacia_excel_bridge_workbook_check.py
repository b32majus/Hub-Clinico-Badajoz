#!/usr/bin/env python3
"""Check the Cáceres FH Excel Bridge workbook and its TSV truth boundary."""

from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

import generate_farmacia_excel_bridge_workbook as generator


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "templates" / "PROMueve_FH_Caceres_Bridge_DEMO.xlsx"
EXPECTED_SHEETS = [name for name, _ in generator.OPERATIONAL_SHEETS] + list(generator.TECHNICAL_SHEETS)
EXPECTED_TABLES = dict(generator.OPERATIONAL_SHEETS)
EXPECTED_PROCESSOR_COUNTS = {
    "PATIENTS": 4,
    "REQUESTS": 4,
    "VALIDATIONS": 4,
    "TREATMENTS": 4,
    "TREATMENT_LINES": 4,
    "TREATMENT_MOVEMENTS": 2,
    "VISITS": 4,
    "VISIT_LINES": 8,
    "OBSERVATIONS": 12,
    "PROMS": 4,
    "ADVERSE_EVENTS": 2,
    "CAUSALITY": 4,
    "AUDIT_EVENTS": 8,
    "IMPORT_ERRORS": 0,
    "RENEWAL_CYCLES": 0,
    "RENEWAL_TASKS": 0,
}


class CheckFailure(AssertionError):
    def __init__(self, code: str, message: str) -> None:
        super().__init__(message)
        self.code = code


def require(condition: bool, message: str, code: str = "FAIL_STRUCTURE") -> None:
    if not condition:
        raise CheckFailure(code, message)


def load_core_truth() -> dict:
    script = r"""
const fs = require('fs');
const path = require('path');
const vm = require('vm');
const root = process.argv[1];
const sandbox = { window: {} };
vm.createContext(sandbox);
vm.runInContext(fs.readFileSync(path.join(root, 'scripts/farmacia_export_v2_core.js'), 'utf8'), sandbox);
const core = sandbox.window.FarmaciaExportV2Core;
const fixture = name => JSON.parse(fs.readFileSync(path.join(root, 'data/demo/farmacia/export_v2', name), 'utf8'));
const validationFixture = fixture('validation_event_v2.json');
const firstFixture = fixture('first_visit_event_v2.json');
const followFixture = fixture('followup_event_v2.json');

function buildService(namespace, label) {
  const patientValidation = `patient-${namespace}-validation`;
  const patientLongitudinal = `patient-${namespace}-longitudinal`;
  const validationEvent = {
    ...validationFixture.event,
    event_id: `evt-${namespace}-validation`,
    source_event_id: `src-${namespace}-validation`,
    patient_id: patientValidation,
    request_id: `request-${namespace}-validation`,
    validation_id: `validation-${namespace}-normal`,
    hospital_code: null,
    service_code: null,
    service_label: label
  };
  const validationRows = core.projectEventRows(validationEvent, [{ rowKey: `row-${namespace}-validation` }]);
  const formulaEvent = {
    ...validationEvent,
    event_id: `evt-${namespace}-formula-neutral`,
    source_event_id: `src-${namespace}-formula-neutral`,
    request_id: `request-${namespace}-formula-neutral`,
    validation_id: `validation-${namespace}-formula-neutral`,
    identifier_value: `CIP-${namespace.toUpperCase()}-VALIDATION`,
    requested_drug_name: '=1+1',
    requested_active_ingredient: '+SUM(A1:A2)',
    requested_presentation: '-2+3',
    requested_dose_text: '@command'
  };
  const formulaRows = core.projectEventRows(formulaEvent, [{ rowKey: `row-${namespace}-formula-neutral` }]);
  const firstEvent = {
    ...firstFixture.event,
    event_id: `evt-${namespace}-first-visit`,
    source_event_id: `src-${namespace}-first-visit`,
    patient_id: patientLongitudinal,
    first_visit_id: `first-visit-${namespace}`,
    hospital_code: null,
    service_code: null,
    service_label: label
  };
  const firstPayloads = [1, 2].map(index => ({
    ...firstFixture.rowPayloads[0],
    rowKey: `row-${namespace}-first-line-${index}`,
    treatment_id: `treatment-${namespace}-${index}`,
    line_id: `line-${namespace}-${index}`,
    line_role: index === 1 ? 'primary' : 'additional',
    is_primary_line: index === 1,
    line_drug_name: `Fármaco sintético ${namespace} ${index}`
  }));
  const firstRows = core.projectEventRows(firstEvent, firstPayloads);
  const followEvent = {
    ...followFixture.event,
    event_id: `evt-${namespace}-followup`,
    source_event_id: `src-${namespace}-followup`,
    patient_id: patientLongitudinal,
    visit_id: `visit-${namespace}-followup`,
    adverse_event_id: `adverse-${namespace}-followup`,
    hospital_code: null,
    service_code: null,
    service_label: label,
    adverse_event_suspects_json: [
      { suspect_ref: `line-${namespace}-1`, reported: true },
      { suspect_ref: `line-${namespace}-2`, reported: false }
    ],
    causality_assessments_json: [
      { suspect_ref: `line-${namespace}-1`, method: 'DEMO_A', score: 0 },
      { suspect_ref: `line-${namespace}-2`, method: 'DEMO_B', result: null }
    ]
  };
  const followPayloads = followFixture.rowPayloads.map((payload, index) => ({
    ...payload,
    rowKey: `row-${namespace}-follow-line-${index + 1}`,
    treatment_id: `treatment-${namespace}-${index + 1}`,
    line_id: `line-${namespace}-${index + 1}`,
    line_drug_name: `Fármaco posterior ${namespace} ${index + 1}`
  }));
  const followRows = core.projectEventRows(followEvent, followPayloads);
  const acts = [validationRows, firstRows, followRows, formulaRows];
  const ordered = acts.flat();
  return {
    namespace,
    patientValidation,
    patientLongitudinal,
    rows: ordered,
    tsv: acts.map(rows => core.serializeRowsToTsv(rows)).join('\n'),
    formulaRowId: formulaRows[0].row_id,
    firstSourceEventId: firstEvent.source_event_id,
    followSourceEventId: followEvent.source_event_id
  };
}

process.stdout.write(JSON.stringify({
  columns: core.ROW_COLUMNS,
  services: {
    '01_DERMA': buildService('der', 'Dermatología'),
    '03_DIGESTIVO': buildService('dig', 'Digestivo')
  }
}));
"""
    result = subprocess.run(
        ["node", "-e", script, str(ROOT)],
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    require(result.returncode == 0, f"export core fixture projection failed: {result.stderr}")
    return json.loads(result.stdout)


def table_values(worksheet) -> list:
    return list(worksheet.tables.values())


def manifest_tables() -> dict[str, dict]:
    return {item["sheet"]: item for item in generator.load_relational_manifest()["technical_tables"]}


def assert_package_has_no_active_content(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        names = [name.lower() for name in archive.namelist()]
    forbidden = ("vbaproject", "externallinks", "connections", "embeddings", "activex")
    require(not any(token in name for name in names for token in forbidden), "workbook contains active or external package parts")


def logical_table_rows(worksheet, definition: dict) -> list[list]:
    table = table_values(worksheet)[0]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows = [[worksheet.cell(row, column).value for column in range(min_col, max_col + 1)] for row in range(min_row + 1, max_row + 1)]
    if len(rows) == 1 and all(value is None for value in rows[0]):
        return []
    require(not any(all(value is None for value in row) for row in rows), f"internal empty row in {definition['sheet']}")
    return rows


def assert_structure(path: Path, *, expect_native_empty: bool, expected_native_rows: int | None = None) -> None:
    require(path.is_file(), f"workbook does not exist: {path}")
    assert_package_has_no_active_content(path)
    workbook = load_workbook(path, data_only=False, keep_links=True)
    require(workbook.sheetnames == EXPECTED_SHEETS, "sheet list or order differs")
    require(not workbook._external_links, "workbook contains external links")
    columns = generator.load_columns()
    for sheet_name, table_name in EXPECTED_TABLES.items():
        worksheet = workbook[sheet_name]
        require(worksheet.sheet_state == "visible", f"unexpected visibility for {sheet_name}")
        tables = table_values(worksheet)
        require(len(tables) == 1 and tables[0].displayName == table_name, f"unexpected input table in {sheet_name}")
        expected_ref = "A1:EV2" if expected_native_rows is None else f"A1:EV{expected_native_rows + 1}"
        require(tables[0].ref == expected_ref, f"unexpected table range in {sheet_name}: {tables[0].ref}")
        require([worksheet.cell(1, index).value for index in range(1, 153)] == columns, f"152-column header parity failed in {sheet_name}")
        require(worksheet.freeze_panes == "A2", f"freeze pane differs in {sheet_name}")
        if expect_native_empty:
            require(all(worksheet.cell(2, index).value is None for index in range(1, 153)), f"{sheet_name} contains a native row")
        for index in range(1, 153):
            require(worksheet.cell(2, index).number_format == "@", f"row 2 is not Text in {sheet_name} column {index}")
            require(worksheet.column_dimensions[worksheet.cell(2, index).column_letter].number_format == "@", f"column is not Text in {sheet_name} column {index}")
    for sheet_name, definition in manifest_tables().items():
        worksheet = workbook[sheet_name]
        require(worksheet.sheet_state == "hidden", f"unexpected visibility for {sheet_name}")
        tables = table_values(worksheet)
        require(len(tables) == 1, f"technical sheet {sheet_name} must contain one table")
        require(tables[0].displayName == definition["table"], f"technical table name differs in {sheet_name}")
        require([cell.value for cell in worksheet[1]] == definition["headers"], f"technical headers differ in {sheet_name}")
        require(worksheet.freeze_panes == "A2", f"technical freeze pane differs in {sheet_name}")
        for column in range(1, len(definition["headers"]) + 1):
            require(worksheet.cell(2, column).number_format == "@", f"technical placeholder is not Text in {sheet_name}")
    require(not any(name.startswith("APP_") for name in workbook.sheetnames), "APP_* sheet found")
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                require(cell.data_type != "f", f"formula found at {worksheet.title}!{cell.coordinate}", "FAIL_FORMULA_DETECTED")


def semantic_snapshot(path: Path) -> dict:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    return {
        "active": workbook.active.title,
        "sheets": [
            {
                "name": worksheet.title,
                "state": worksheet.sheet_state,
                "freeze": str(worksheet.freeze_panes or ""),
                "tables": sorted((table.displayName, table.ref) for table in table_values(worksheet)),
                "values": [[cell.value for cell in row] for row in worksheet.iter_rows()],
                "formats": [[cell.number_format for cell in row] for row in worksheet.iter_rows()],
            }
            for worksheet in workbook.worksheets
        ],
    }


def truth_corpora(core_truth: dict) -> dict[str, str]:
    require(core_truth["columns"] == generator.load_columns(), "core ROW_COLUMNS differs from schema x-column-order")
    corpora = {sheet: service["tsv"] for sheet, service in core_truth["services"].items()}
    require(set(corpora) == set(EXPECTED_TABLES), "truth services differ from workbook inputs")
    for sheet, corpus in corpora.items():
        lines = corpus.split("\n")
        require(len(lines) == 6, f"{sheet} must have six truth rows")
        require(all(len(line.split("\t")) == 152 for line in lines), f"{sheet} truth row is not 152 columns")
    require(corpora["01_DERMA"] != corpora["03_DIGESTIVO"], "Derma and Digestivo truth corpora must differ")
    return corpora


def paste_tsv(worksheet, tsv: str) -> None:
    lines = tsv.split("\n")
    for row_offset, line in enumerate(lines, start=2):
        cells = line.split("\t")
        require(len(cells) == 152, "truth TSV row does not contain 152 cells")
        for column, cell_text in enumerate(cells, start=1):
            cell = worksheet.cell(row_offset, column)
            cell.value = None if cell_text == "" else cell_text
            cell.number_format = "@"
    table_values(worksheet)[0].ref = f"A1:EV{len(lines) + 1}"


def reconstruct_tsv(worksheet, row_count: int) -> str:
    return "\n".join(
        "\t".join("" if worksheet.cell(row, column).value is None else str(worksheet.cell(row, column).value) for column in range(1, 153))
        for row in range(2, row_count + 2)
    )


def assert_qa_rows(workbook, corpora: dict[str, str]) -> None:
    for sheet_name, corpus in corpora.items():
        worksheet = workbook[sheet_name]
        expected_rows = [line.split("\t") for line in corpus.split("\n")]
        for row_offset, expected_row in enumerate(expected_rows, start=2):
            for column, expected_value in enumerate(expected_row, start=1):
                cell = worksheet.cell(row_offset, column)
                require(cell.data_type != "f", f"formula found at {sheet_name}!{cell.coordinate}", "FAIL_FORMULA_DETECTED")
                if expected_value == "":
                    require(cell.value is None, f"expected empty cell changed at {sheet_name}!{cell.coordinate}", "FAIL_TSV_ROUNDTRIP_MISMATCH")
                else:
                    require(isinstance(cell.value, str) and cell.number_format == "@", f"value type or Text format changed at {sheet_name}!{cell.coordinate}", "FAIL_VALUE_OR_TYPE_MISMATCH")
        require(reconstruct_tsv(worksheet, 6).encode() == corpus.encode(), f"manual Excel TSV differs in {sheet_name}", "FAIL_TSV_ROUNDTRIP_MISMATCH")


def assert_formula_neutrality(workbook, core_truth: dict, processed: bool = False) -> None:
    columns = generator.load_columns()
    expected = {
        "requested_drug_name": '"=1+1"',
        "requested_active_ingredient": '"+SUM(A1:A2)"',
        "requested_presentation": '"-2+3"',
        "requested_dose_text": '"@command"',
    }
    for sheet_name, service in core_truth["services"].items():
        row_ids = [json.loads(line.split("\t")[columns.index("row_id")]) for line in service["tsv"].split("\n")]
        formula_row = row_ids.index(service["formulaRowId"]) + 2
        for field, canonical in expected.items():
            cell = workbook[sheet_name].cell(formula_row, columns.index(field) + 1)
            require(cell.value == canonical, f"formula-neutral text changed in {sheet_name} for {field}")
            require(cell.data_type != "f", f"formula created in {sheet_name} for {field}")
        if processed:
            require(workbook[sheet_name].cell(formula_row, 1).value == '"PROCESADA"', f"formula-neutral row not processed in {sheet_name}")


def run_truth_test(candidate: Path, core_truth: dict) -> None:
    corpora = truth_corpora(core_truth)
    with tempfile.TemporaryDirectory(prefix="fh-excel-bridge-truth-") as temp_dir:
        path = Path(temp_dir) / "truth.xlsx"
        workbook = load_workbook(candidate, data_only=False, keep_links=True)
        for sheet_name, corpus in corpora.items():
            paste_tsv(workbook[sheet_name], corpus)
        workbook.save(path)
        reopened = load_workbook(path, data_only=False, keep_links=True)
        assert_qa_rows(reopened, corpora)
        assert_formula_neutrality(reopened, core_truth)


def run_negative_controls(candidate: Path, core_truth: dict) -> list[str]:
    corpora = truth_corpora(core_truth)
    controls = (
        ("FAIL_FORMULA_DETECTED", lambda worksheet: setattr(worksheet["AM7"], "value", "=1+1")),
        ("FAIL_VALUE_OR_TYPE_MISMATCH", lambda worksheet: setattr(worksheet["K2"], "value", 1)),
        ("FAIL_TSV_ROUNDTRIP_MISMATCH", lambda worksheet: setattr(worksheet["A2"], "value", worksheet["E2"].value)),
    )
    passed: list[str] = []
    with tempfile.TemporaryDirectory(prefix="fh-excel-bridge-negative-") as temp_dir:
        for index, (expected, mutate) in enumerate(controls, start=1):
            path = Path(temp_dir) / f"negative-{index}.xlsx"
            workbook = load_workbook(candidate, data_only=False, keep_links=True)
            for sheet_name, corpus in corpora.items():
                paste_tsv(workbook[sheet_name], corpus)
            mutate(workbook["01_DERMA"])
            workbook.save(path)
            try:
                assert_structure(path, expect_native_empty=False, expected_native_rows=6)
                checked = load_workbook(path, data_only=False, keep_links=True)
                assert_qa_rows(checked, corpora)
                assert_formula_neutrality(checked, core_truth)
            except CheckFailure as error:
                require(error.code == expected, f"negative control returned {error.code}, expected {expected}")
                passed.append(expected)
                continue
            raise CheckFailure("FAIL_NEGATIVE_CONTROL", f"negative control did not produce {expected}")
    return passed


def decode_technical(text) -> object:
    if text is None or text == "":
        return None
    if text == "TRUE":
        return True
    if text == "FALSE":
        return False
    return json.loads(text)


def check_processor_qa_copy(path: Path, core_truth: dict) -> tuple[dict[str, str], list[str]]:
    assert_structure(path, expect_native_empty=False, expected_native_rows=6)
    workbook = load_workbook(path, data_only=False, keep_links=True)
    corpora = truth_corpora(core_truth)
    columns = generator.load_columns()
    mutable = {columns.index(field) for field in ("bridge_status", "bridge_processed_at", "bridge_error_code", "bridge_error_detail")}
    for sheet_name, corpus in corpora.items():
        expected_rows = [line.split("\t") for line in corpus.split("\n")]
        worksheet = workbook[sheet_name]
        for row_offset, expected in enumerate(expected_rows, start=2):
            for column_index, expected_text in enumerate(expected):
                actual = worksheet.cell(row_offset, column_index + 1).value
                if column_index in mutable:
                    continue
                require(("" if actual is None else str(actual)) == expected_text, f"processor changed raw field in {sheet_name} row {row_offset}")
            require(worksheet.cell(row_offset, 1).value == '"PROCESADA"', f"row not PROCESADA in {sheet_name} row {row_offset}")
    assert_formula_neutrality(workbook, core_truth, processed=True)
    definitions = manifest_tables()
    for sheet_name, expected_count in EXPECTED_PROCESSOR_COUNTS.items():
        rows = logical_table_rows(workbook[sheet_name], definitions[sheet_name])
        require(len(rows) == expected_count, f"{sheet_name} has {len(rows)} rows, expected {expected_count}")
    treatment_headers = definitions["TREATMENTS"]["headers"]
    treatment_rows = logical_table_rows(workbook["TREATMENTS"], definitions["TREATMENTS"])
    source_index = treatment_headers.index("first_seen_source_event_id")
    expected_first_sources = {service["firstSourceEventId"] for service in core_truth["services"].values()}
    require({decode_technical(row[source_index]) for row in treatment_rows} == expected_first_sources, "treatment first_seen source changed after follow-up")
    line_headers = definitions["TREATMENT_LINES"]["headers"]
    line_rows = logical_table_rows(workbook["TREATMENT_LINES"], definitions["TREATMENT_LINES"])
    line_source_index = line_headers.index("first_seen_source_event_id")
    require({decode_technical(row[line_source_index]) for row in line_rows} == expected_first_sources, "line first_seen source changed after follow-up")
    return corpora, []


def check_clean_candidate(path: Path, core_truth: dict) -> tuple[dict[str, str], list[str]]:
    assert_structure(path, expect_native_empty=True)
    with tempfile.TemporaryDirectory(prefix="fh-excel-bridge-generated-") as temp_dir:
        generated = generator.build_workbook(Path(temp_dir) / "generated.xlsx")
        assert_structure(generated, expect_native_empty=True)
        require(semantic_snapshot(path) == semantic_snapshot(generated), "versioned workbook differs semantically from clean generation")
    run_truth_test(path, core_truth)
    return truth_corpora(core_truth), run_negative_controls(path, core_truth)


def check_manual_qa_copy(path: Path, core_truth: dict) -> tuple[dict[str, str], list[str]]:
    corpora = truth_corpora(core_truth)
    assert_structure(path, expect_native_empty=False, expected_native_rows=6)
    workbook = load_workbook(path, data_only=False, keep_links=True)
    assert_qa_rows(workbook, corpora)
    assert_formula_neutrality(workbook, core_truth)
    for sheet_name, definition in manifest_tables().items():
        require(not logical_table_rows(workbook[sheet_name], definition), f"technical table {sheet_name} populated before processor")
    return corpora, run_negative_controls(DEFAULT_WORKBOOK, core_truth)


def emit_manual_tsv(path: Path, corpora: dict[str, str]) -> list[Path]:
    digestivo = path.with_name(f"{path.stem}_digestivo{path.suffix or '.tsv'}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(corpora["01_DERMA"], encoding="utf-8", newline="")
    digestivo.write_text(corpora["03_DIGESTIVO"], encoding="utf-8", newline="")
    return [path, digestivo]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    modes = parser.add_mutually_exclusive_group()
    modes.add_argument("--qa-copy", action="store_true", help="validate an Excel-saved copy containing the raw truth corpus")
    modes.add_argument("--processor-qa-copy", action="store_true", help="validate an Excel-saved copy after Office Script processing")
    modes.add_argument("--emit-manual-tsv", type=Path, help="write differentiated Derma and Digestivo TSV corpora")
    args = parser.parse_args()
    core_truth = load_core_truth()
    corpora = truth_corpora(core_truth)
    if args.emit_manual_tsv:
        paths = emit_manual_tsv(args.emit_manual_tsv, corpora)
        print("\n".join(f"manual_tsv={path}" for path in paths))
        return
    if args.processor_qa_copy:
        corpora, negative_controls = check_processor_qa_copy(args.workbook, core_truth)
        mode = "processor_excel_qa"
    elif args.qa_copy:
        corpora, negative_controls = check_manual_qa_copy(args.workbook, core_truth)
        mode = "manual_excel_qa"
    else:
        corpora, negative_controls = check_clean_candidate(args.workbook, core_truth)
        mode = "candidate"
    truth_bytes = sum(len(corpus.encode("utf-8")) for corpus in corpora.values())
    print("farmacia_excel_bridge_workbook_check: PASS")
    print(f"mode={mode} sheets={len(EXPECTED_SHEETS)} columns=152 truth_rows=12 truth_bytes={truth_bytes}")
    print(f"negative_controls={','.join(negative_controls)}")


if __name__ == "__main__":
    try:
        main()
    except (AssertionError, ValueError, json.JSONDecodeError, subprocess.CalledProcessError) as error:
        code = error.code if isinstance(error, CheckFailure) else "FAIL_UNEXPECTED"
        print(f"farmacia_excel_bridge_workbook_check: {code}: {error}", file=sys.stderr)
        raise SystemExit(1)
