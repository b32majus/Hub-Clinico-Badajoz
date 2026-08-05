#!/usr/bin/env python3
"""Generate the Cáceres FH Excel Bridge workbook container."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SCHEMA_PATH = ROOT / "schemas" / "farmacia_export_row_v2.schema.json"
RELATIONAL_MANIFEST_PATH = ROOT / "schemas" / "farmacia_excel_bridge_relational_v1.json"
DEFAULT_OUTPUT = ROOT / "templates" / "PROMueve_FH_Caceres_Bridge_DEMO.xlsx"

OPERATIONAL_SHEETS = (
    ("01_DERMA", "tblBridgeDermaInput"),
    ("03_DIGESTIVO", "tblBridgeDigestivoInput"),
)
def load_columns() -> list[str]:
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    columns = schema.get("x-column-order")
    if not isinstance(columns, list) or len(columns) != 152:
        raise ValueError("x-column-order must contain exactly 152 columns")
    if any(not isinstance(name, str) or not name for name in columns):
        raise ValueError("x-column-order must contain non-empty string names")
    if len(set(columns)) != 152:
        raise ValueError("x-column-order must contain 152 unique names")
    return columns


def load_relational_manifest() -> dict:
    manifest = json.loads(RELATIONAL_MANIFEST_PATH.read_text(encoding="utf-8"))
    if manifest.get("manifest_version") != "1.0.0-draft.1":
        raise ValueError("unexpected relational manifest version")
    if manifest.get("input_column_count") != 152:
        raise ValueError("relational manifest must require 152 input columns")
    expected_inputs = [
        {"sheet": sheet_name, "table": table_name}
        for sheet_name, table_name in OPERATIONAL_SHEETS
    ]
    if manifest.get("input_tables") != expected_inputs:
        raise ValueError("relational manifest input tables differ")
    tables = manifest.get("technical_tables")
    if not isinstance(tables, list) or len(tables) != 16:
        raise ValueError("relational manifest must define 16 technical tables")
    names: set[str] = set()
    table_names: set[str] = set()
    for item in tables:
        if not isinstance(item, dict):
            raise ValueError("technical table definition must be an object")
        sheet = item.get("sheet")
        table = item.get("table")
        headers = item.get("headers")
        key = item.get("key")
        if not isinstance(sheet, str) or not sheet or sheet in names:
            raise ValueError("technical sheet names must be unique non-empty strings")
        if not isinstance(table, str) or not table or table in table_names:
            raise ValueError("technical table names must be unique non-empty strings")
        if not isinstance(headers, list) or not headers or any(not isinstance(value, str) or not value for value in headers):
            raise ValueError(f"technical table {sheet} has invalid headers")
        if len(headers) != len(set(headers)):
            raise ValueError(f"technical table {sheet} has duplicate headers")
        if not isinstance(key, list) or not key or any(value not in headers for value in key):
            raise ValueError(f"technical table {sheet} has invalid key")
        names.add(sheet)
        table_names.add(table)
    if '"movement_type"' in RELATIONAL_MANIFEST_PATH.read_text(encoding="utf-8"):
        raise ValueError("non-canonical movement_type alias found")
    return manifest


TECHNICAL_TABLES = tuple(load_relational_manifest()["technical_tables"])
TECHNICAL_SHEETS = tuple(item["sheet"] for item in TECHNICAL_TABLES)


def build_workbook(output: Path = DEFAULT_OUTPUT) -> Path:
    columns = load_columns()
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.creator = "PROMueve"
    workbook.title = "PROMueve FH Cáceres Excel Bridge DEMO"
    workbook.subject = "Excel Bridge V4 workbook container"
    workbook.description = "Synthetic/demo workbook container; no real patient data."

    table_style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, table_name in OPERATIONAL_SHEETS:
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        for index, name in enumerate(columns, start=1):
            header = worksheet.cell(row=1, column=index, value=name)
            header.fill = header_fill
            header.font = header_font
            header.alignment = Alignment(horizontal="center", vertical="center")
            placeholder = worksheet.cell(row=2, column=index)
            placeholder.number_format = "@"
            column_letter = placeholder.column_letter
            worksheet.column_dimensions[column_letter].number_format = "@"
            worksheet.column_dimensions[column_letter].width = max(12, min(28, len(name) + 2))
        worksheet.row_dimensions[1].height = 30
        table = Table(displayName=table_name, ref="A1:EV2")
        table.tableStyleInfo = table_style
        worksheet.add_table(table)

    for definition in TECHNICAL_TABLES:
        sheet_name = definition["sheet"]
        headers = definition["headers"]
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.sheet_state = "hidden"
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        for index, name in enumerate(headers, start=1):
            header = worksheet.cell(row=1, column=index, value=name)
            header.fill = header_fill
            header.font = header_font
            header.alignment = Alignment(horizontal="center", vertical="center")
            placeholder = worksheet.cell(row=2, column=index)
            placeholder.number_format = "@"
            column_letter = placeholder.column_letter
            worksheet.column_dimensions[column_letter].number_format = "@"
            worksheet.column_dimensions[column_letter].width = max(12, min(28, len(name) + 2))
        worksheet.row_dimensions[1].height = 30
        last_column = get_column_letter(len(headers))
        table = Table(displayName=definition["table"], ref=f"A1:{last_column}2")
        table.tableStyleInfo = table_style
        worksheet.add_table(table)

    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    output = build_workbook(args.output)
    print(output)


if __name__ == "__main__":
    main()
