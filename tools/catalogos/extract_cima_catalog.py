#!/usr/bin/env python3
"""
extract_cima_catalog.py — Extracción del catálogo farmacológico desde CIMA/AEMPS
================================================================================

Objetivo:
  Generar un Excel con el catálogo dual (CIMA oficial + local manual)
  para el Hub Clínico Badajoz, módulo de Farmacia Hospitalaria.

Fuente:
  API REST pública de CIMA/AEMPS (https://cima.aemps.es/cima/rest/)
  Documentación: CIMA REST API v1.23

Uso:
  python3 extract_cima_catalog.py

Output:
  /srv/kairos-lab/outbox/exports/catalogo-cima-v0-1/hub_catalogo_farmacologico_dual_cima_local_v0_1_REAL_YYYYMMDD.xlsx

Restricciones:
  - No requiere autenticación
  - No usa datos reales de pacientes
  - Sin conexión a GitHub, systemd, ni servicios del VPS
"""

import json
import time
import logging
import sys
import os
import re
from datetime import datetime, timezone
from urllib.request import urlopen, Request
from urllib.error import HTTPError, URLError

# ─── Configuración ───────────────────────────────────────────────────────────

BASE_URL = "https://cima.aemps.es/cima/rest"
PAGE_SIZE = 200          # Máximo permitido por CIMA
MAX_PAGES = 200          # Límite de seguridad (~40.000 registros)
REQUEST_DELAY = 0.15     # Segundos entre peticiones (respetar servidor)
TIMEOUT_SEC = 60         # Timeout por petición HTTP

OUTPUT_DIR = "/srv/kairos-lab/outbox/exports/catalogo-cima-v0-1"
os.makedirs(OUTPUT_DIR, exist_ok=True)

TODAY = datetime.now(timezone.utc).strftime("%Y%m%d")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"hub_catalogo_farmacologico_dual_cima_local_v0_1_REAL_{TODAY}.xlsx")

# ─── Logging ─────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(os.path.join(OUTPUT_DIR, f"extract_log_{TODAY}.txt"), mode="w")
    ]
)
log = logging.getLogger("cima-extract")

# ─── Utilidades HTTP ─────────────────────────────────────────────────────────

def api_get(endpoint: str, params: dict = None, retries: int = 3) -> dict:
    """Realiza una GET a la API de CIMA con reintentos."""
    url = f"{BASE_URL}/{endpoint}"
    if params:
        qs = "&".join(f"{k}={v}" for k, v in params.items() if v is not None)
        url = f"{url}?{qs}"

    for attempt in range(retries):
        try:
            req = Request(url, headers={"Accept": "application/json",
                                        "User-Agent": "KairOS-Hub-Clinico-Badajoz/1.0"})
            with urlopen(req, timeout=TIMEOUT_SEC) as resp:
                data = json.loads(resp.read().decode("utf-8"))
            time.sleep(REQUEST_DELAY)
            return data
        except HTTPError as e:
            if e.code == 429:
                wait = 5 * (attempt + 1)
                log.warning(f"HTTP 429 (rate limit). Esperando {wait}s...")
                time.sleep(wait)
                continue
            log.error(f"HTTP {e.code} en {url}: {e.reason}")
            raise
        except (URLError, OSError) as e:
            if attempt < retries - 1:
                wait = 3 * (attempt + 1)
                log.warning(f"Error de conexión (intento {attempt+1}/{retries}): {e}. Reintentando en {wait}s...")
                time.sleep(wait)
                continue
            log.error(f"Error de conexión tras {retries} intentos: {e}")
            raise
        except json.JSONDecodeError as e:
            log.error(f"Error decodificando JSON: {e}")
            raise

    log.error(f"Fallaron todos los reintentos para {url}")
    return {}

# ─── Extracción de medicamentos ──────────────────────────────────────────────

def extract_all_medicamentos() -> list:
    """
    Obtiene todos los medicamentos comercializados desde CIMA.
    Retorna lista de diccionarios con datos planos.
    """
    all_records = []
    total_filas = None
    sync_log = {
        "started_at": datetime.now(timezone.utc).isoformat(),
        "source": "CIMA/AEMPS REST API",
        "endpoint_used": f"{BASE_URL}/medicamentos",
        "records_retrieved": 0,
        "records_written": 0,
        "records_failed": 0,
        "warnings": [],
        "status": "running",
        "notes": ""
    }

    for page in range(1, MAX_PAGES + 1):
        try:
            resp = api_get("medicamentos", {
                "comerc": 1,
                "pagina": page,
                "tamano": PAGE_SIZE
            })

            if total_filas is None:
                total_filas = resp.get("totalFilas", 0)
                total_pages = (total_filas + PAGE_SIZE - 1) // PAGE_SIZE
                log.info(f"Total registros CIMA: {total_filas} (~{total_pages} páginas)")

            resultados = resp.get("resultados", [])
            if not resultados:
                log.info(f"Página {page}: sin resultados. Fin de datos.")
                break

            # Procesar cada resultado
            for med in resultados:
                try:
                    record = process_medicamento(med)
                    all_records.append(record)
                except Exception as e:
                    sync_log["records_failed"] += 1
                    sync_log["warnings"].append(f"Error procesando nregistro={med.get('nregistro','?')}: {e}")
                    continue

            sync_log["records_retrieved"] += len(resultados)
            log.info(f"Página {page}: {len(resultados)} registros (total acumulado: {sync_log['records_retrieved']})")

            # Si recibimos menos registros que el tamaño de página, terminamos
            if len(resultados) < PAGE_SIZE:
                log.info(f"Página {page}: menos registros que tamaño de página. Fin de datos.")
                break

        except Exception as e:
            sync_log["warnings"].append(f"Error en página {page}: {e}")
            log.error(f"Error en página {page}: {e}")
            # Si falla una página, continuamos con la siguiente
            continue

    sync_log["finished_at"] = datetime.now(timezone.utc).isoformat()
    sync_log["records_written"] = len(all_records)
    sync_log["status"] = "completed" if all_records else "failed"

    log.info(f"Extracción completada: {len(all_records)} registros")
    return all_records, sync_log


def process_medicamento(med: dict) -> dict:
    """Procesa un medicamento de CIMA y devuelve un registro plano."""

    # Datos base
    nregistro = med.get("nregistro", "")
    nombre = med.get("nombre", "")
    labtitular = med.get("labtitular", "")
    cpresc_raw = med.get("cpresc", "")
    comerc = med.get("comerc", False)
    receta = med.get("receta", False)
    generico = med.get("generico", False)
    biosimilar = med.get("biosimilar", False)
    dosis = med.get("dosis", "")

    # Forma farmacéutica
    forma_farmaceutica = ""
    ff = med.get("formaFarmaceutica")
    if ff and isinstance(ff, dict):
        forma_farmaceutica = ff.get("nombre", "")

    # Vías de administración
    vias = []
    for via in med.get("viasAdministracion", []):
        if isinstance(via, dict) and via.get("nombre"):
            vias.append(via["nombre"])
    via_str = "; ".join(vias)

    # Principios activos (texto plano)
    pactivos_raw = ""
    principios = med.get("principiosActivos", [])
    if principios:
        pa_parts = []
        for pa in principios:
            if isinstance(pa, dict):
                nombre_pa = pa.get("nombre", "")
                cantidad = pa.get("cantidad", "")
                unidad = pa.get("unidad", "")
                if cantidad:
                    pa_parts.append(f"{nombre_pa} {cantidad}{unidad}")
                else:
                    pa_parts.append(nombre_pa)
        pactivos_raw = ", ".join(pa_parts)

    # ATC codes
    atc_codes = []
    for atc in med.get("atcs", []):
        if isinstance(atc, dict) and atc.get("codigo"):
            atc_codes.append(atc["codigo"])
    atc_str = "; ".join(atc_codes)

    # Documentos (FT + Prospecto)
    url_ft = ""
    url_prospecto = ""
    for doc in med.get("docs", []):
        if isinstance(doc, dict):
            if doc.get("tipo") == 1:
                url_ft = doc.get("url", "") or doc.get("urlHtml", "")
            elif doc.get("tipo") == 2:
                url_prospecto = doc.get("url", "") or doc.get("urlHtml", "")

    # Estado
    estado = med.get("estado", {})
    fecha_aut = ""
    if isinstance(estado, dict):
        aut_ts = estado.get("aut")
        if aut_ts:
            try:
                fecha_aut = datetime.fromtimestamp(aut_ts / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
            except (OSError, ValueError):
                fecha_aut = str(aut_ts)

    # Derivar nombre comercial desde el nombre
    # En CIMA, el nombre del medicamento suele ser el nombre comercial + forma
    nombre_comercial = ""
    if nombre:
        # Intentar extraer nombre comercial (parte antes de la dosis/forma)
        # Ej: "ASPIRINA C 400 mg/240 mg COMPRIMIDOS EFERVESCENTES"
        # Buscar patrón de dosis para separar
        match = re.match(r'^(.+?)\s+\d+\s*(?:mg|UI|mcg|g|ml)/', nombre)
        if match:
            nombre_comercial = match.group(1).strip()
        else:
            match = re.match(r'^(.+?)\s+\d+\s*(?:mg|UI|mcg|g|ml)\s', nombre)
            if match:
                nombre_comercial = match.group(1).strip()
            else:
                nombre_comercial = nombre

    # ─── Filtro hospitalario derivado ────────────────────────────────────────
    es_hospitalario = False
    criterio_hospitalario = ""
    revisar = False

    cpresc_upper = cpresc_raw.upper() if cpresc_raw else ""

    # Patrones que indican uso/ámbito hospitalario
    patrones_hospitalarios = [
        ("H", "Contiene 'H' en condiciones de prescripción"),
        ("DH", "Contiene 'DH' (Diagnóstico Hospitalario)"),
        ("UH", "Contiene 'UH' (Uso Hospitalario)"),
        ("HOSPITALARIO", "Contiene literal 'HOSPITALARIO'"),
        ("HOSPITAL", "Contiene literal 'HOSPITAL'"),
        ("ÁMBITO HOSPITALARIO", "Contiene 'ÁMBITO HOSPITALARIO'"),
        ("AMBITO HOSPITALARIO", "Contiene 'AMBITO HOSPITALARIO'"),
        ("USO HOSPITALARIO", "Contiene 'USO HOSPITALARIO'"),
        ("DIAGNÓSTICO HOSPITALARIO", "Contiene 'DIAGNÓSTICO HOSPITALARIO'"),
        ("DISPENSACIÓN HOSPITALARIA", "Contiene 'DISPENSACIÓN HOSPITALARIA'"),
        ("HOSPITALES", "Contiene 'HOSPITALES'"),
    ]

    for patron, descripcion in patrones_hospitalarios:
        if patron in cpresc_upper:
            if patron == "H" and len(patron) == 1:
                # 'H' sola puede aparecer en muchos contextos
                # Solo marcar si aparece como etiqueta separada
                if re.search(r'\bH\b', cpresc_upper):
                    es_hospitalario = True
                    criterio_hospitalario = descripcion
                    break
                elif "H " in cpresc_upper or " H" in cpresc_upper or cpresc_upper == "H":
                    es_hospitalario = True
                    criterio_hospitalario = descripcion
                    break
            else:
                es_hospitalario = True
                criterio_hospitalario = descripcion
                break

    # Si no se encontró patrón claro pero el medicamento podría ser hospitalario
    # por su naturaleza (biológicos, oncológicos, etc.), marcamos como revisar
    if not es_hospitalario and biosimilar:
        es_hospitalario = None  # None = "revisar"
        criterio_hospitalario = "Biosimilar - requiere revisión manual"

    # Normalizar el campo es_hospitalario_derivado
    if es_hospitalario is True:
        es_hosp_str = "TRUE"
    elif es_hospitalario is None:
        es_hosp_str = "revisar"
    else:
        es_hosp_str = "FALSE"

    # Fecha de actualización del medicamento en CIMA
    # Usamos fecha de los documentos como proxy de actualización
    fecha_actualizacion = ""
    for doc in med.get("docs", []):
        if isinstance(doc, dict) and doc.get("fecha"):
            try:
                f = datetime.fromtimestamp(doc["fecha"] / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
                if not fecha_actualizacion or f > fecha_actualizacion:
                    fecha_actualizacion = f
            except (OSError, ValueError):
                pass

    return {
        "drug_source_id": nregistro,
        "source_type": "CIMA",
        "codigo_nacional": "",
        "nregistro": nregistro,
        "nombre_presentacion": nombre,
        "nombre_comercial": nombre_comercial,
        "principio_activo": pactivos_raw,
        "forma_farmaceutica": forma_farmaceutica,
        "dosis_presentacion": dosis,
        "via": via_str,
        "laboratorio": labtitular,
        "comercializado": str(comerc),
        "receta": str(receta),
        "cpresc_raw": cpresc_raw,
        "cpresc_normalizado": normalizar_cpresc(cpresc_raw),
        "es_hospitalario_derivado": es_hosp_str,
        "criterio_hospitalario": criterio_hospitalario,
        "biosimilar": str(biosimilar),
        "generico": str(generico),
        "problema_suministro": "",
        "atc_codes": atc_str,
        "url_ficha_tecnica": url_ft,
        "url_prospecto": url_prospecto,
        "fecha_autorizacion": fecha_aut,
        "fecha_actualizacion_cima": fecha_actualizacion,
        "fecha_extraccion": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "activo_en_catalogo": "TRUE",
        "observaciones_import": ""
    }


def normalizar_cpresc(cpresc_raw: str) -> str:
    """Normaliza las condiciones de prescripción a categorías estándar."""
    if not cpresc_raw:
        return ""
    upper = cpresc_raw.upper()
    if "SIN RECETA" in upper:
        return "Sin Receta"
    if "MEDICAMENTO SUJETO A PRESCRIPCIÓN MÉDICA" in upper:
        if "TRATAMIENTO DE LARGA DURACIÓN" in upper:
            return "Prescripción Médica (Larga Duración)"
        return "Prescripción Médica"
    if "DIAGNÓSTICO HOSPITALARIO" in upper or "DH" in cpresc_raw:
        return "Diagnóstico Hospitalario"
    if "USO HOSPITALARIO" in upper:
        return "Uso Hospitalario"
    if "MEDICAMENTO PUBLICITARIO" in upper:
        return "Medicamento Publicitario"
    return cpresc_raw[:100]


def get_problemas_suministro() -> dict:
    """Obtiene problemas de suministro activos de CIMA."""
    try:
        resp = api_get("psuministro")
        problemas = {}
        for item in resp if isinstance(resp, list) else []:
            if isinstance(item, dict) and item.get("cn"):
                problemas[item["cn"]] = {
                    "activo": item.get("activo", False),
                    "observ": item.get("observ", ""),
                    "tipo": item.get("tipoProblemaSuministro", "")
                }
        log.info(f"Problemas de suministro obtenidos: {len(problemas)}")
        return problemas
    except Exception as e:
        log.warning(f"No se pudieron obtener problemas de suministro: {e}")
        return {}


def get_presentaciones_for_medicamento(nregistro: str) -> list:
    """Obtiene las presentaciones de un medicamento (para CN)."""
    try:
        resp = api_get("presentaciones", {"nregistro": nregistro})
        return resp.get("resultados", [])
    except Exception:
        return []


# ─── Generación Excel ────────────────────────────────────────────────────────

def generate_excel(records: list, sync_log: dict, problemas: dict):
    """Genera el Excel con todas las hojas requeridas."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log.error("openpyxl no está instalado. Instala con: uv pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    today_dt = datetime.now(timezone.utc)

    # Estilos
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="003B5C", end_color="003B5C", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    def write_sheet(ws, title, headers, rows=None, col_widths=None):
        """Escribe una hoja con cabecera estilizada."""
        ws.title = title
        # Cabecera
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        # Datos
        if rows:
            for row_idx, row in enumerate(rows, 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
                    cell.alignment = cell_alignment
                    cell.border = thin_border
        # Ajustar ancho de columnas
        if col_widths:
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + col_idx)].width = width
        else:
            ws.column_dimensions["A"].width = 15

    # ─── Hoja 1: CATALOGO_CIMA ──────────────────────────────────────────────
    ws1 = wb.active
    headers_cima = [
        "drug_source_id", "source_type", "codigo_nacional", "nregistro",
        "nombre_presentacion", "nombre_comercial", "principio_activo",
        "forma_farmaceutica", "dosis_presentacion", "via", "laboratorio",
        "comercializado", "receta", "cpresc_raw", "cpresc_normalizado",
        "es_hospitalario_derivado", "criterio_hospitalario",
        "biosimilar", "generico", "problema_suministro",
        "atc_codes", "url_ficha_tecnica", "url_prospecto",
        "fecha_autorizacion", "fecha_actualizacion_cima", "fecha_extraccion",
        "activo_en_catalogo", "observaciones_import"
    ]
    rows_cima = []
    for r in records:
        rows_cima.append([r.get(h, "") for h in headers_cima])
    write_sheet(ws1, "CATALOGO_CIMA", headers_cima, rows_cima)
    log.info(f"Hoja CATALOGO_CIMA: {len(rows_cima)} registros")

    # ─── Hoja 2: CATALOGO_LOCAL_ESPECIAL ────────────────────────────────────
    ws2 = wb.create_sheet()
    headers_local = [
        "local_drug_id", "source_type", "display_name", "nombre_comercial_si_existe",
        "principio_activo_o_molecula", "codigo_desarrollo", "presentacion_texto",
        "via", "forma_farmaceutica", "dosis_texto", "tipo_situacion",
        "estado_desarrollo", "origen", "pais_origen_si_aplica",
        "requiere_validacion_fh", "estado_validacion_fh", "fecha_alta_local",
        "fecha_ultima_revision", "responsable_revision", "observaciones",
        "activo_en_catalogo", "reconciliado_con_cima_id"
    ]
    # Filas de ejemplo DEMO (sintéticas, no reales)
    local_ejemplos = [
        ["LOC-DEMO-001", "LOCAL_ESPECIAL", "Comprimido experimental X", "",
         "Molécula X en investigación", "MOL-X-001", "Comp 10mg",
         "Oral", "Comprimido", "10 mg", "ENSAYO_CLINICO",
         "EN_DESARROLLO", "Protocolo local", "",
         "SÍ", "PENDIENTE", "2026-06-01", "2026-06-06",
         "Pendiente asignar", "DEMO: Ejemplo de uso compasivo", "SÍ", ""],
        ["LOC-DEMO-002", "LOCAL_ESPECIAL", "Medicamento extranjero Y", "",
         "Fármaco Y importado", "EXT-Y", "Inyectable 50mg/ml",
         "SC", "Solución inyectable", "50 mg/ml", "MEDICACION_EXTRANJERA",
         "AUTORIZADO_USO_COMPASIVO", "Importación", "Francia",
         "SÍ", "VALIDADO", "2026-05-15", "2026-06-01",
         "Dr. García (Farmacia)", "DEMO: Ejemplo medicación extranjera", "SÍ", ""],
    ]
    write_sheet(ws2, "CATALOGO_LOCAL_ESPECIAL", headers_local, local_ejemplos)

    # ─── Hoja 3: CATALOGO_ALIAS ─────────────────────────────────────────────
    ws3 = wb.create_sheet()
    headers_alias = ["alias", "drug_id", "source_type", "alias_type", "activo"]
    alias_rows = []
    # Generar alias básicos desde los primeros 500 registros (para no alargar)
    for r in records[:500]:
        nreg = r.get("nregistro", "")
        nombre = r.get("nombre_presentacion", "")
        nombre_com = r.get("nombre_comercial", "")
        pa = r.get("principio_activo", "")
        if nombre:
            alias_rows.append([nombre, nreg, "CIMA", "nombre_presentacion", "SÍ"])
        if nombre_com and nombre_com != nombre:
            alias_rows.append([nombre_com, nreg, "CIMA", "nombre_comercial", "SÍ"])
        if pa:
            alias_rows.append([pa, nreg, "CIMA", "principio_activo", "SÍ"])
    write_sheet(ws3, "CATALOGO_ALIAS", headers_alias, alias_rows)

    # ─── Hoja 4: CATALOGO_FAVORITOS_CIRCUITO ────────────────────────────────
    ws4 = wb.create_sheet()
    headers_fav = ["drug_id", "source_type", "servicio", "patologia", "circuito",
                   "prioridad", "activo", "observaciones"]
    write_sheet(ws4, "CATALOGO_FAVORITOS_CIRCUITO", headers_fav, [])

    # ─── Hoja 5: TRATAMIENTOS_PACIENTE ──────────────────────────────────────
    ws5 = wb.create_sheet()
    headers_tx = ["treatment_id", "patient_id", "selected_drug_id",
                  "selected_drug_source_type", "nombre_snapshot",
                  "principio_activo_snapshot", "presentacion_snapshot",
                  "via_snapshot", "fecha_inicio", "fecha_fin",
                  "dosis_prescrita", "pauta_prescrita", "patologia_contexto",
                  "indicacion_texto", "tipo_uso", "estado_validacion_fh",
                  "observaciones_fh", "activo", "created_at", "updated_at"]
    write_sheet(ws5, "TRATAMIENTOS_PACIENTE", headers_tx, [])

    # ─── Hoja 6: SYNC_LOG ──────────────────────────────────────────────────
    ws6 = wb.create_sheet()
    headers_sync = ["sync_id", "started_at", "finished_at", "source",
                    "endpoint_used", "records_retrieved", "records_written",
                    "records_failed", "warnings", "status", "notes"]
    sync_rows = [[
        f"SYNC-{TODAY}-001",
        sync_log.get("started_at", ""),
        sync_log.get("finished_at", ""),
        sync_log.get("source", ""),
        sync_log.get("endpoint_used", ""),
        sync_log.get("records_retrieved", 0),
        sync_log.get("records_written", 0),
        sync_log.get("records_failed", 0),
        "\n".join(sync_log.get("warnings", [])),
        sync_log.get("status", ""),
        f"Extracción de catálogo CIMA v0.1 para Hub Clínico Badajoz. Problemas de suministro obtenidos: {len(problemas)}"
    ]]
    write_sheet(ws6, "SYNC_LOG", headers_sync, sync_rows)

    # ─── Hoja 7: VALIDACIONES ──────────────────────────────────────────────
    ws7 = wb.create_sheet()
    headers_val = ["validacion", "resultado", "detalle"]

    sin_cn = sum(1 for r in records if not r.get("codigo_nacional", ""))
    sin_pa = sum(1 for r in records if not r.get("principio_activo", "").strip())
    sin_nombre = sum(1 for r in records if not r.get("nombre_presentacion", "").strip())
    hospitalarios = sum(1 for r in records if r.get("es_hospitalario_derivado") == "TRUE")
    revisar_hosp = sum(1 for r in records if r.get("es_hospitalario_derivado") == "revisar")
    con_receta = sum(1 for r in records if r.get("receta") == "True")
    biosimilares = sum(1 for r in records if r.get("biosimilar") == "True")

    # Duplicados por nregistro
    nregistros = [r.get("nregistro", "") for r in records]
    duplicados_nreg = len(nregistros) - len(set(filter(None, nregistros)))

    validation_rows = [
        ["Número total de registros CIMA", len(records),
         "Registros obtenidos del endpoint medicamentos?comerc=1"],
        ["Registros con nregistro", len(records) - sum(1 for r in records if not r.get("nregistro", "")),
         ""],
        ["Registros con código nacional", len(records) - sin_cn,
         "El endpoint medicamentos no devuelve CN; está vacío en esta versión. Las presentaciones tienen CN."],
        ["Registros sin código nacional", sin_cn,
         "Pendiente de obtener presentaciones para completar CN"],
        ["Registros con principio activo vacío", sin_pa,
         "Medicamentos sin principios activos estructurados"],
        ["Registros con nombre vacío", sin_nombre, ""],
        ["Duplicados por nregistro", duplicados_nreg, ""],
        ["Marcado como hospitalario derivado", hospitalarios,
         "Basado en condiciones de prescripción (cpresc)"],
        ["Marcado como revisar hospitalario", revisar_hosp,
         "Biosimilares sin patrón hospitalario claro - requieren revisión manual de Farmacia"],
        ["Criterios usados para hospitalario",
         "Patrones en cpresc: H, DH, UH, HOSPITALARIO, USO HOSPITALARIO, DIAGNÓSTICO HOSPITALARIO, DISPENSACIÓN HOSPITALARIA",
         "Filtro conservador: no elimina registros, solo marca columna derivada"],
        ["Con receta médica", con_receta,
         "campo receta=True en CIMA"],
        ["Biosimilares", biosimilares,
         ""],
        ["Errores de parsing", sync_log.get("records_failed", 0),
         "Registros que fallaron durante el procesamiento"],
        ["Warnings", len(sync_log.get("warnings", [])),
         "\n".join(sync_log.get("warnings", [])) or "Ninguno"],
        ["Timestamp extracción", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"), ""],
    ]
    write_sheet(ws7, "VALIDACIONES", headers_val, validation_rows)

    # ─── Hoja 8: LISTAS ────────────────────────────────────────────────────
    ws8 = wb.create_sheet()
    headers_listas = ["lista", "valor", "descripcion"]
    listas_rows = [
        ["tipo_situacion", "ENSAYO_CLINICO", "Medicamento en ensayo clínico"],
        ["tipo_situacion", "USO_COMPASIVO", "Uso compasivo autorizado"],
        ["tipo_situacion", "MEDICACION_EXTRANJERA", "Medicamento no autorizado en España, importado"],
        ["tipo_situacion", "PRECOMERCIALIZACION", "Fármaco antes de autorización comercial"],
        ["tipo_situacion", "SIN_NOMBRE_COMERCIAL", "Molécula sin nombre comercial asignado"],
        ["tipo_situacion", "FUERA_FICHA_TECNICA", "Uso fuera de indicación autorizada"],
        ["tipo_situacion", "PROTOCOLO_LOCAL", "Medicación según protocolo local del centro"],
        ["tipo_situacion", "OTRA", "Otra situación especial no categorizada"],
        ["", "", ""],
        ["estado_desarrollo", "EN_DESARROLLO", ""],
        ["estado_desarrollo", "AUTORIZADO_USO_COMPASIVO", ""],
        ["estado_desarrollo", "APROBADO_COMITÉ", ""],
        ["estado_desarrollo", "EN_EVALUACION", ""],
        ["estado_desarrollo", "SUSPENDIDO", ""],
        ["estado_desarrollo", "RETIRADO", ""],
        ["", "", ""],
        ["estado_validacion_fh", "PENDIENTE", "Pendiente de validación por Farmacia"],
        ["estado_validacion_fh", "VALIDADO", "Validado por Farmacia"],
        ["estado_validacion_fh", "RECHAZADO", "Rechazado por Farmacia"],
        ["estado_validacion_fh", "EN_REVISION", "En revisión por Farmacia"],
        ["", "", ""],
        ["source_type", "CIMA", "Procedente del catálogo oficial CIMA/AEMPS"],
        ["source_type", "LOCAL_ESPECIAL", "Registrado manualmente por Farmacia (situaciones especiales)"],
        ["source_type", "LOCAL_FAVORITO", "Favorito/marcado por circuito clínico"],
        ["", "", ""],
        ["activo", "TRUE", "Activo en el catálogo"],
        ["activo", "FALSE", "Desactivado (no se muestra en búsquedas)"],
        ["", "", ""],
        ["sí/no", "SÍ", ""],
        ["sí/no", "NO", ""],
    ]
    write_sheet(ws8, "LISTAS", headers_listas, listas_rows)

    # Guardar
    wb.save(OUTPUT_FILE)
    log.info(f"Excel guardado: {OUTPUT_FILE}")
    return OUTPUT_FILE


# ─── Main ─────────────────────────────────────────────────────────────────────

def enrich_with_presentaciones(records: list, problemas: dict) -> list:
    """
    Obtiene presentaciones de CIMA y enriquece los registros con:
    - código nacional (cn)
    - principios activos (pactivos)
    - problemas de suministro
    
    Retorna los registros enriquecidos.
    """
    log.info("\n--- Enriqueciendo con datos de presentaciones CIMA ---")
    
    # Construir lookup de presentaciones por nregistro
    pres_cache_file = os.path.join(OUTPUT_DIR, f"cima_presentaciones_{TODAY}.json")
    
    if os.path.exists(pres_cache_file):
        log.info(f"Cache de presentaciones encontrada: {pres_cache_file}")
        with open(pres_cache_file, "r") as f:
            presentaciones_por_nreg = json.load(f)
    else:
        presentaciones_por_nreg = {}
        page = 1
        while page <= MAX_PAGES:
            try:
                resp = api_get("presentaciones", {
                    "comerc": 1,
                    "pagina": page,
                    "tamano": PAGE_SIZE
                })
                resultados = resp.get("resultados", [])
                if not resultados:
                    break
                for pres in resultados:
                    nreg = pres.get("nregistro", "")
                    if nreg and nreg not in presentaciones_por_nreg:
                        presentaciones_por_nreg[nreg] = {
                            "cn": pres.get("cn", ""),
                            "pactivos": pres.get("pactivos", ""),
                        }
                log.info(f"Presentaciones página {page}: {len(resultados)} registros (nregistros únicos: {len(presentaciones_por_nreg)})")
                if len(resultados) < PAGE_SIZE:
                    break
                page += 1
            except Exception as e:
                log.warning(f"Error obteniendo presentaciones página {page}: {e}")
                break
        
        # Guardar cache
        with open(pres_cache_file, "w") as f:
            json.dump(presentaciones_por_nreg, f, ensure_ascii=False)
        log.info(f"Presentaciones guardadas en caché: {pres_cache_file}")
    
    log.info(f"Total nregistros con presentaciones: {len(presentaciones_por_nreg)}")
    
    # Enriquecer registros
    enriched = 0
    for record in records:
        nreg = record.get("nregistro", "")
        if nreg in presentaciones_por_nreg:
            pres = presentaciones_por_nreg[nreg]
            if not record.get("codigo_nacional"):
                record["codigo_nacional"] = pres.get("cn", "")
            if not record.get("principio_activo"):
                record["principio_activo"] = pres.get("pactivos", "")
            enriched += 1
    
    # También enriquecer con problemas de suministro
    cn_problemas = {}
    for cn_val, info in problemas.items():
        cn_problemas[cn_val] = info.get("observ", "")
    
    # Para registros con CN, marcar problemas de suministro
    for record in records:
        cn = record.get("codigo_nacional", "")
        if cn and cn in cn_problemas:
            record["problema_suministro"] = cn_problemas[cn]
    
    log.info(f"Registros enriquecidos: {enriched}/{len(records)}")
    return records


def main():
    log.info("=" * 60)
    log.info("EXTRACCIÓN CATÁLOGO CIMA v0.1")
    log.info(f"Fecha: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}")
    log.info("=" * 60)
    log.info(f"URL base API: {BASE_URL}")
    log.info(f"Output dir: {OUTPUT_DIR}")

    # Paso 1: Obtener problemas de suministro (rápido, antes que el resto)
    log.info("\n--- Obteniendo problemas de suministro activos ---")
    problemas = get_problemas_suministro()

    # Paso 2: Extraer todos los medicamentos paginados
    log.info("\n--- Extrayendo medicamentos comercializados ---")
    cache_file = os.path.join(OUTPUT_DIR, f"cima_raw_data_{TODAY}.json")
    
    # Intentar cargar desde caché primero
    if os.path.exists(cache_file):
        log.info(f"Cache encontrada: {cache_file}. Cargando datos guardados...")
        with open(cache_file, "r") as f:
            cached = json.load(f)
        records = cached.get("records", [])
        sync_log = cached.get("sync_log", {})
        log.info(f"Cargados {len(records)} registros desde caché")
    else:
        records, sync_log = extract_all_medicamentos()
        if records:
            with open(cache_file, "w") as f:
                json.dump({"records": records, "sync_log": sync_log}, f, ensure_ascii=False, indent=2)
            log.info(f"Datos guardados en caché: {cache_file}")

    if not records:
        log.error("No se obtuvieron registros. Abortando.")
        sync_log["status"] = "failed"
        sync_log["notes"] = "Sin registros obtenidos. Posible error de conexión o API."
        generate_excel([], sync_log, problemas)
        print(f"\nOUTPUT_FILE={OUTPUT_FILE}")
        print(f"STATUS=failed")
        sys.exit(1)

    # Paso 3: Enriquecer con presentaciones (CN + principios activos)
    records = enrich_with_presentaciones(records, problemas)

    # Actualizar sync_log con datos enriquecidos
    sync_log["notes"] = f"Enriquecido con presentaciones. {sum(1 for r in records if r.get('codigo_nacional',''))} registros con CN, {sum(1 for r in records if r.get('principio_activo',''))} con principios activos."

    # Paso 4: Generar Excel
    log.info(f"\n--- Generando Excel con {len(records)} registros ---")
    excel_path = generate_excel(records, sync_log, problemas)

    # Estadísticas finales
    hospitalarios = sum(1 for r in records if r.get("es_hospitalario_derivado") == "TRUE")
    revisar_hosp = sum(1 for r in records if r.get("es_hospitalario_derivado") == "revisar")
    biosimilares = sum(1 for r in records if r.get("biosimilar") == "True")

    print(f"\n{'='*60}")
    print(f"EXTRACCIÓN COMPLETADA")
    print(f"{'='*60}")
    print(f"Registros totales CIMA:    {len(records)}")
    print(f"Biosimilares:              {biosimilares}")
    print(f"Hospitalarios (derivado):  {hospitalarios}")
    print(f"Revisar (dudosos):         {revisar_hosp}")
    print(f"Problemas suministro:      {len(problemas)}")
    print(f"Errores:                   {sync_log.get('records_failed', 0)}")
    print(f"Warnings:                  {len(sync_log.get('warnings', []))}")
    print(f"")
    print(f"Excel: {excel_path}")
    print(f"Status: completed" if sync_log['status'] == 'completed' else f"Status: {sync_log['status']}")
    print(f"{'='*60}")

    return excel_path, records, sync_log


if __name__ == "__main__":
    main()
