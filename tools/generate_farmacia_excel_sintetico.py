#!/usr/bin/env python3
"""WO8.1c — Generate synthetic FH Excel populated with stress-test data.

Creates templates/farmacia_excel_operativo_FH_WO8_v1_sintetico.xlsx
with 40+ synthetic rows across 4 service sheets, plus catalog of special drugs.

All data are clearly synthetic — no real patient data, no real CIP/NHC.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, sys, json
from datetime import datetime, date, timedelta
import random

random.seed(42)

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(REPO_ROOT, "templates", "farmacia_excel_operativo_FH_WO8_v1.xlsx")
OUTPUT_PATH = os.path.join(REPO_ROOT, "templates", "farmacia_excel_operativo_FH_WO8_v1_sintetico.xlsx")

# ── Column headers (61 cols, blocks A-H) ──────────────────────────
SERVICE_COLS = [
    # A
    "patient_id", "cip_demo_o_hash", "nhc_o_codigo_interno",
    "fecha_nacimiento_o_edad", "sexo", "servicio_origen", "patologia_indicacion",
    # B
    "fecha_acto", "tipo_acto_fh", "visita_id", "validacion_id",
    "tratamiento_id", "linea_id", "profesional_fh", "estado_registro",
    # C
    "marca_comercial", "principio_activo", "codigo_nacional",
    "numero_registro", "source_type", "categoria_farmaco",
    "tipo_relacion", "estado_linea", "tipo_movimiento", "es_principal",
    "fecha_inicio", "fecha_fin", "motivo_inicio_cambio_suspension",
    # D
    "dosis_presentacion", "via", "pauta_codigo", "pauta_label", "pauta_otro_texto",
    # E
    "tipo_validacion", "resultado_validacion", "requiere_prebiologico",
    "tb_estado", "serologias_estado", "vacunas_estado",
    "bloqueantes_validacion", "observaciones_validacion",
    # F
    "adherencia_morisky", "haq", "eva_dolor", "dlqi",
    "respuesta_clinica", "incidencias", "observaciones_seguimiento",
    # G
    "hay_efecto_adverso", "ea_id", "ea_descripcion", "ea_gravedad",
    "farmaco_sospechoso_id", "farmaco_sospechoso_nombre",
    "causalidad_naranjo", "causalidad_karch", "accion_ea",
    # H
    "created_at", "updated_at", "demo_flag", "observaciones_generales",
]

HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SERVICES = {
    "01_DERMA": {"full": "Dermatología", "prefix": "DER"},
    "02_REUMA": {"full": "Reumatología", "prefix": "REU"},
    "03_DIGESTIVO": {"full": "Digestivo", "prefix": "DIG"},
    "04_ONCO": {"full": "Oncología", "prefix": "ONC"},
}

NOW = datetime(2026, 6, 14, 10, 0, 0)
ISO_NOW = NOW.strftime("%Y-%m-%d %H:%M:%S")

# ── Common drug catalog (brand · INN) ──────────────────────────────
DRUGS_DERMA = [
    ("Cosentyx", "Secukinumab"),
    ("Humira", "Adalimumab"),
    ("Bimzelx", "Bimekizumab"),
    ("Taltz", "Ixekizumab"),
    ("Skyrizi", "Risankizumab"),
]
DRUGS_REUMA = [
    ("Orencia", "Abatacept"),
    ("Benlysta", "Belimumab"),
    ("Rixathon", "Rituximab"),
    ("RoActemra", "Tocilizumab"),
    ("Olumiant", "Baricitinib"),
]
DRUGS_DIGESTIVO = [
    ("Stelara", "Ustekinumab"),
    ("Entyvio", "Vedolizumab"),
    ("Remicade", "Infliximab"),
    ("Skyrizi", "Risankizumab"),
    ("Humira", "Adalimumab"),
]
DRUGS_ONCO = [
    ("Keytruda", "Pembrolizumab"),
    ("Opdivo", "Nivolumab"),
    ("Tecentriq", "Atezolizumab"),
    ("Avastin", "Bevacizumab"),
    ("Herceptin", "Trastuzumab"),
]

PAUTAS = [
    ("SC-2SEM", "Subcutánea cada 2 semanas", "SC 2 semanas"),
    ("SC-4SEM", "Subcutánea cada 4 semanas", "SC 4 semanas"),
    ("IV-4SEM", "Intravenosa cada 4 semanas", "IV 4 semanas"),
    ("IV-8SEM", "Intravenosa cada 8 semanas", "IV 8 semanas"),
    ("SC-SEM", "Subcutánea semanal", "SC semanal"),
    ("IV-2SEM", "Intravenosa cada 2 semanas", "IV 2 semanas"),
]

PROFESIONALES = [
    "Dr. García (FH)",
    "Dra. López (FH)",
    "Dr. Martínez (FH)",
    "Dra. Sánchez (FH)",
]


def make_row(service, sidx, tipo_acto, extra=None):
    """Build a 61-element row array."""
    prefix = SERVICES[service]["prefix"]
    pnum = sidx + 1
    pat_id = f"FH-SYN-{prefix}-{pnum:03d}"
    cip = f"DEMO-CIP-{prefix}-{pnum:03d}"
    nhc = f"NHC-SYN-{prefix}-{pnum:04d}"
    sexo = random.choice(["M", "F"])
    patologia = extra.get("patologia", "")
    fecha_acto = extra.get("fecha_acto", "2026-06-01")
    visita_id = extra.get("visita_id", "")
    validacion_id = extra.get("validacion_id", "")
    tratamiento_id = extra.get("tratamiento_id", "")
    linea_id = extra.get("linea_id", "")
    profesional = extra.get("profesional", random.choice(PROFESIONALES))
    marca = extra.get("marca_comercial", "")
    principio = extra.get("principio_activo", "")
    cod_nac = extra.get("codigo_nacional", "")
    nreg = extra.get("numero_registro", "")
    source = extra.get("source_type", "DEMO")
    cat_farm = extra.get("categoria_farmaco", "biológico")
    tipo_rel = extra.get("tipo_relacion", "principal")
    est_linea = extra.get("estado_linea", "activo")
    tipo_mov = extra.get("tipo_movimiento", "")
    es_principal = extra.get("es_principal", "TRUE")
    fecha_ini = extra.get("fecha_inicio", "")
    fecha_fin = extra.get("fecha_fin", "")
    motivo = extra.get("motivo", "")
    dosis = extra.get("dosis", "")
    via = extra.get("via", "")
    pauta_cod = extra.get("pauta_codigo", "")
    pauta_label = extra.get("pauta_label", "")
    pauta_otro = extra.get("pauta_otro_texto", "")
    tipo_val = extra.get("tipo_validacion", "")
    res_val = extra.get("resultado_validacion", "")
    req_pre = extra.get("requiere_prebiologico", "")
    tb = extra.get("tb_estado", "")
    sero = extra.get("serologias_estado", "")
    vac = extra.get("vacunas_estado", "")
    bloq = extra.get("bloqueantes_validacion", "")
    obs_val = extra.get("observaciones_validacion", "")
    adherencia = extra.get("adherencia_morisky", "")
    haq = extra.get("haq", "")
    eva = extra.get("eva_dolor", "")
    dlqi = extra.get("dlqi", "")
    resp_clin = extra.get("respuesta_clinica", "")
    incidencias = extra.get("incidencias", "")
    obs_seg = extra.get("observaciones_seguimiento", "")
    hay_ea = extra.get("hay_efecto_adverso", "FALSE")
    ea_id = extra.get("ea_id", "")
    ea_desc = extra.get("ea_descripcion", "")
    ea_grav = extra.get("ea_gravedad", "")
    fs_id = extra.get("farmaco_sospechoso_id", "")
    fs_nom = extra.get("farmaco_sospechoso_nombre", "")
    naranjo = extra.get("causalidad_naranjo", "")
    karch = extra.get("causalidad_karch", "")
    acc_ea = extra.get("accion_ea", "")
    obs_gen = extra.get("observaciones_generales", "")

    row = [
        pat_id, cip, nhc, str(random.randint(25, 75)), sexo,
        SERVICES[service]["full"], patologia,
        fecha_acto, tipo_acto, visita_id, validacion_id,
        tratamiento_id, linea_id, profesional, "completado",
        marca, principio, cod_nac, nreg, source, cat_farm,
        tipo_rel, est_linea, tipo_mov, es_principal,
        fecha_ini, fecha_fin, motivo,
        dosis, via, pauta_cod, pauta_label, pauta_otro,
        tipo_val, res_val, req_pre, tb, sero, vac, bloq, obs_val,
        adherencia, haq, eva, dlqi, resp_clin, incidencias, obs_seg,
        hay_ea, ea_id, ea_desc, ea_grav, fs_id, fs_nom, naranjo, karch, acc_ea,
        ISO_NOW, ISO_NOW, "TRUE", obs_gen,
    ]
    return row


# ── Build all rows per service ────────────────────────────────────

def build_derma_rows():
    """10+ rows for Dermatología covering required case types."""
    rows = []
    # Helper to pick drug
    def drug(i):
        b, p = DRUGS_DERMA[i]
        return b, p

    # 1. HS pendiente de validación inicial
    rows.append(make_row("01_DERMA", 0, "validacion_inicial", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-05-28",
        "marca_comercial": "Cosentyx", "principio_activo": "Secukinumab",
        "tipo_validacion": "inicial", "resultado_validacion": "pendiente",
        "requiere_prebiologico": "TRUE",
        "tb_estado": "pendiente", "serologias_estado": "pendiente",
        "vacunas_estado": "pendiente",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-28",
        "dosis": "300 mg", "via": "SC",
        "observaciones_validacion": "Pendiente completar estudio prebiológico",
        "observaciones_generales": "HS pendiente de validación inicial",
    }))

    # 2. HS con validación inicial completada
    rows.append(make_row("01_DERMA", 1, "validacion_inicial", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-05-15",
        "marca_comercial": "Humira", "principio_activo": "Adalimumab",
        "codigo_nacional": "123456", "numero_registro": "REG-ADL-001",
        "tipo_validacion": "inicial", "resultado_validacion": "validado",
        "requiere_prebiologico": "TRUE",
        "tb_estado": "negativo", "serologias_estado": "completo", "vacunas_estado": "completo",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-20",
        "dosis": "40 mg", "via": "SC",
        "pauta_codigo": "SC-2SEM", "pauta_label": "Subcutánea cada 2 semanas",
        "observaciones_validacion": "Validación inicial completada sin incidencias",
    }))

    # 3. HS en primera visita
    rows.append(make_row("01_DERMA", 2, "primera_visita", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-06-01",
        "marca_comercial": "Humira", "principio_activo": "Adalimumab",
        "tipo_validacion": "inicial", "resultado_validacion": "validado",
        "tratamiento_id": "TRAT-HS-002", "linea_id": "L-HS-002",
        "visita_id": "VIS-HS-002-01", "validacion_id": "VAL-HS-002",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-20",
        "dosis": "40 mg", "via": "SC",
        "pauta_codigo": "SC-2SEM", "pauta_label": "Subcutánea cada 2 semanas",
        "adherencia_morisky": "Alta", "eva_dolor": "3", "dlqi": "8",
        "respuesta_clinica": "Mejoría parcial de lesiones",
        "observaciones_seguimiento": "Primera visita tras validación. Buena tolerancia.",
    }))

    # 4. HS en seguimiento rutinario
    rows.append(make_row("01_DERMA", 3, "seguimiento", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-06-10",
        "marca_comercial": "Humira", "principio_activo": "Adalimumab",
        "tratamiento_id": "TRAT-HS-002", "linea_id": "L-HS-002",
        "visita_id": "VIS-HS-002-02",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-20",
        "dosis": "40 mg", "via": "SC",
        "pauta_codigo": "SC-2SEM", "pauta_label": "Subcutánea cada 2 semanas",
        "adherencia_morisky": "Alta", "haq": "1.2", "eva_dolor": "2", "dlqi": "5",
        "respuesta_clinica": "Estable, reducción de abscesos",
    }))

    # 5. HS con efecto adverso leve
    rows.append(make_row("01_DERMA", 4, "seguimiento", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-06-08",
        "marca_comercial": "Bimzelx", "principio_activo": "Bimekizumab",
        "tratamiento_id": "TRAT-HS-005", "linea_id": "L-HS-005",
        "visita_id": "VIS-HS-005-01",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-10",
        "dosis": "320 mg", "via": "SC",
        "pauta_codigo": "SC-4SEM", "pauta_label": "Subcutánea cada 4 semanas",
        "adherencia_morisky": "Alta",
        "hay_efecto_adverso": "TRUE",
        "ea_id": "EA-HS-005-01",
        "ea_descripcion": "Candidiasis oral leve",
        "ea_gravedad": "leve",
        "farmaco_sospechoso_id": "TRAT-HS-005",
        "farmaco_sospechoso_nombre": "Bimzelx (Bimekizumab)",
        "causalidad_naranjo": "posible",
        "causalidad_karch": "posible",
        "accion_ea": "Tratamiento antifúngico tópico, continuar Bimzelx",
        "observaciones_seguimiento": "EA leve manejado con antifúngico",
        "observaciones_generales": "Candidiasis oral leve secundaria a Bimekizumab",
    }))

    # 6. HS con suspensión
    rows.append(make_row("01_DERMA", 5, "suspension", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-04-15",
        "marca_comercial": "Taltz", "principio_activo": "Ixekizumab",
        "tratamiento_id": "TRAT-HS-006", "linea_id": "L-HS-006",
        "tipo_relacion": "principal", "estado_linea": "suspendido", "es_principal": "FALSE",
        "tipo_movimiento": "suspension",
        "fecha_inicio": "2026-02-01", "fecha_fin": "2026-04-15",
        "motivo_inicio_cambio_suspension": "Falta de respuesta clínica tras 10 semanas",
        "dosis": "80 mg", "via": "SC",
        "pauta_codigo": "SC-2SEM", "pauta_label": "Subcutánea cada 2 semanas",
        "observaciones_validacion": "Suspensión por falta de eficacia",
        "observaciones_generales": "HS sin respuesta a Ixekizumab, pendiente de nueva alternativa",
    }))

    # 7. HS con cambio terapéutico
    rows.append(make_row("01_DERMA", 6, "nueva_validacion_cambio", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-04-20",
        "marca_comercial": "Cosentyx", "principio_activo": "Secukinumab",
        "validacion_id": "VAL-HS-007",
        "tratamiento_id": "TRAT-HS-007", "linea_id": "L-HS-007",
        "tipo_validacion": "cambio", "resultado_validacion": "validado",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "tipo_movimiento": "cambio_terapeutico",
        "fecha_inicio": "2026-04-20",
        "motivo_inicio_cambio_suspension": "Cambio desde Ixekizumab por falta de respuesta",
        "dosis": "300 mg", "via": "SC",
        "pauta_codigo": "SC-4SEM", "pauta_label": "Subcutánea cada 4 semanas",
        "observaciones_validacion": "Cambio a Secukinumab tras fallo a Ixekizumab",
        "observaciones_generales": "Cambio terapéutico: Taltz → Cosentyx",
    }))

    # 8. HS con nueva validación por adición
    rows.append(make_row("01_DERMA", 7, "nueva_validacion_adicion", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-05-25",
        "marca_comercial": "Bimzelx", "principio_activo": "Bimekizumab",
        "validacion_id": "VAL-HS-008",
        "tratamiento_id": "TRAT-HS-008", "linea_id": "L-HS-008",
        "tipo_validacion": "adicion", "resultado_validacion": "validado",
        "tipo_relacion": "adicional", "estado_linea": "anadido", "es_principal": "FALSE",
        "tipo_movimiento": "tratamiento_anadido",
        "fecha_inicio": "2026-06-01",
        "motivo_inicio_cambio_suspension": "Adición de Bimekizumab por respuesta parcial a Cosentyx",
        "dosis": "320 mg", "via": "SC",
        "pauta_codigo": "SC-4SEM", "pauta_label": "Subcutánea cada 4 semanas",
        "observaciones_validacion": "Adición de Bimekizumab como tratamiento complementario",
        "observaciones_generales": "Adición de Bimzelx a Cosentyx por respuesta parcial",
    }))

    # 9. HS con pauta modificada
    rows.append(make_row("01_DERMA", 8, "seguimiento", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-06-05",
        "marca_comercial": "Cosentyx", "principio_activo": "Secukinumab",
        "tratamiento_id": "TRAT-HS-007", "linea_id": "L-HS-007",
        "visita_id": "VIS-HS-007-01",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-04-20",
        "dosis": "300 mg", "via": "SC",
        "pauta_codigo": "SC-2SEM", "pauta_label": "Subcutánea cada 2 semanas",
        "pauta_otro_texto": "Cambio de pauta por respuesta subóptima: de 4 sem a 2 sem",
        "adherencia_morisky": "Alta", "eva_dolor": "4", "dlqi": "10",
        "respuesta_clinica": "Mejoría parcial, se ajusta pauta",
        "incidencias": "Pauta modificada de 4 a 2 semanas por respuesta subóptima",
        "observaciones_generales": "Cambio de pauta: Cosentyx 300 mg SC cada 4→2 semanas",
    }))

    # 10. HS con tratamiento histórico conservado
    rows.append(make_row("01_DERMA", 9, "seguimiento", {
        "patologia": "Hidrosadenitis supurativa (HS)",
        "fecha_acto": "2026-06-12",
        "marca_comercial": "Humira", "principio_activo": "Adalimumab",
        "tratamiento_id": "TRAT-HS-HIST-001", "linea_id": "L-HS-HIST-001",
        "tipo_relacion": "historico", "estado_linea": "historico", "es_principal": "FALSE",
        "fecha_inicio": "2025-11-01", "fecha_fin": "2026-02-28",
        "motivo_inicio_cambio_suspension": "Pérdida de respuesta secundaria",
        "dosis": "40 mg", "via": "SC",
        "pauta_codigo": "SC-SEM", "pauta_label": "Subcutánea semanal",
        "observaciones_generales": "Histórico: Adalimumab suspendido por pérdida de respuesta. Actualmente en Cosentyx.",
    }))

    return rows


def build_reuma_rows():
    """10+ rows for Reumatología."""
    rows = []

    # 1. Paciente con biológico activo
    rows.append(make_row("02_REUMA", 0, "seguimiento", {
        "patologia": "Artritis reumatoide",
        "fecha_acto": "2026-06-10",
        "marca_comercial": "Orencia", "principio_activo": "Abatacept",
        "tratamiento_id": "TRAT-AR-001", "linea_id": "L-AR-001",
        "visita_id": "VIS-AR-001-05",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2025-03-01",
        "dosis": "125 mg", "via": "SC",
        "pauta_codigo": "SC-SEM", "pauta_label": "Subcutánea semanal",
        "adherencia_morisky": "Alta", "haq": "0.8", "eva_dolor": "2", "dlqi": "3",
        "respuesta_clinica": "Estable, buena respuesta mantenida",
    }))

    # 2. Paciente con histórico suspendido
    rows.append(make_row("02_REUMA", 1, "seguimiento", {
        "patologia": "Artritis reumatoide",
        "fecha_acto": "2026-06-10",
        "marca_comercial": "Olumiant", "principio_activo": "Baricitinib",
        "tratamiento_id": "TRAT-AR-HIST-001", "linea_id": "L-AR-HIST-001",
        "visita_id": "VIS-AR-002-01",
        "tipo_relacion": "historico", "estado_linea": "suspendido", "es_principal": "FALSE",
        "tipo_movimiento": "suspension",
        "fecha_inicio": "2025-06-01", "fecha_fin": "2026-01-15",
        "motivo_inicio_cambio_suspension": "Intolerancia gastrointestinal",
        "dosis": "4 mg", "via": "oral",
        "pauta_codigo": "", "pauta_label": "", "pauta_otro_texto": "1 comprimido al día",
        "observaciones_generales": "Baricitinib suspendido por intolerancia. Actualmente en Orencia.",
    }))

    # 3. Cambio Abatacept → Belimumab
    rows.append(make_row("02_REUMA", 2, "nueva_validacion_cambio", {
        "patologia": "LES con afectación articular",
        "fecha_acto": "2026-05-20",
        "marca_comercial": "Benlysta", "principio_activo": "Belimumab",
        "validacion_id": "VAL-REU-003",
        "tratamiento_id": "TRAT-AR-003", "linea_id": "L-AR-003",
        "tipo_validacion": "cambio", "resultado_validacion": "validado",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "tipo_movimiento": "cambio_terapeutico",
        "fecha_inicio": "2026-05-25",
        "motivo_inicio_cambio_suspension": "Cambio desde Abatacept por diagnóstico de LES",
        "dosis": "200 mg", "via": "SC",
        "pauta_codigo": "SC-SEM", "pauta_label": "Subcutánea semanal",
        "observaciones_validacion": "Cambio a Belimumab por progresión a LES",
        "observaciones_generales": "Cambio terapéutico: Orencia (Abatacept) → Benlysta (Belimumab)",
    }))

    # 4. Cambio previo Abatacept (histórico)
    rows.append(make_row("02_REUMA", 2, "suspension", {
        "patologia": "LES con afectación articular",
        "fecha_acto": "2026-05-15",
        "marca_comercial": "Orencia", "principio_activo": "Abatacept",
        "tratamiento_id": "TRAT-AR-HIST-002", "linea_id": "L-AR-HIST-002",
        "tipo_relacion": "historico", "estado_linea": "suspendido", "es_principal": "FALSE",
        "tipo_movimiento": "cambio_terapeutico",
        "fecha_inicio": "2025-09-01", "fecha_fin": "2026-05-15",
        "motivo_inicio_cambio_suspension": "Diagnóstico de LES, cambio a Belimumab",
        "dosis": "125 mg", "via": "SC",
        "pauta_codigo": "SC-SEM", "pauta_label": "Subcutánea semanal",
        "observaciones_generales": "Histórico: Abatacept suspendido por cambio a Belimumab",
    }))

    # 5. Benlysta activo
    rows.append(make_row("02_REUMA", 3, "seguimiento", {
        "patologia": "LES con afectación articular",
        "fecha_acto": "2026-06-12",
        "marca_comercial": "Benlysta", "principio_activo": "Belimumab",
        "tratamiento_id": "TRAT-AR-003", "linea_id": "L-AR-003",
        "visita_id": "VIS-AR-003-01",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-25",
        "dosis": "200 mg", "via": "SC",
        "pauta_codigo": "SC-SEM", "pauta_label": "Subcutánea semanal",
        "adherencia_morisky": "Media", "haq": "1.5", "eva_dolor": "5", "dlqi": "12",
        "respuesta_clinica": "Mejoría parcial del lupus articular",
        "observaciones_generales": "Benlysta activo, respuesta parcial",
    }))

    # 6. Rixathon añadido
    rows.append(make_row("02_REUMA", 4, "nueva_validacion_adicion", {
        "patologia": "Artritis reumatoide seropositiva",
        "fecha_acto": "2026-06-01",
        "marca_comercial": "Rixathon", "principio_activo": "Rituximab",
        "validacion_id": "VAL-AR-004",
        "tratamiento_id": "TRAT-AR-004", "linea_id": "L-AR-004",
        "tipo_validacion": "adicion", "resultado_validacion": "validado",
        "tipo_relacion": "adicional", "estado_linea": "anadido", "es_principal": "FALSE",
        "tipo_movimiento": "tratamiento_anadido",
        "fecha_inicio": "2026-06-05",
        "motivo_inicio_cambio_suspension": "Adición por respuesta parcial a Abatacept",
        "dosis": "1000 mg", "via": "IV",
        "pauta_codigo": "IV-2SEM", "pauta_label": "Intravenosa cada 2 semanas",
        "observaciones_validacion": "Adición de Rituximab por respuesta articular insuficiente",
    }))

    # 7. Varios fármacos activos (Benlysta + Rituximab ambas activas)
    rows.append(make_row("02_REUMA", 4, "seguimiento", {
        "patologia": "Artritis reumatoide seropositiva",
        "fecha_acto": "2026-06-12",
        "marca_comercial": "Rixathon", "principio_activo": "Rituximab",
        "tratamiento_id": "TRAT-AR-004", "linea_id": "L-AR-004",
        "visita_id": "VIS-AR-004-01",
        "tipo_relacion": "adicional", "estado_linea": "activo", "es_principal": "FALSE",
        "fecha_inicio": "2026-06-05",
        "dosis": "1000 mg", "via": "IV",
        "pauta_codigo": "IV-2SEM", "pauta_label": "Intravenosa cada 2 semanas",
        "adherencia_morisky": "Alta",
        "respuesta_clinica": "Mejoría articular significativa desde adición",
    }))

    # 8. Sospechoso de EA asociado a fármaco
    rows.append(make_row("02_REUMA", 5, "seguimiento", {
        "patologia": "Artritis reumatoide",
        "fecha_acto": "2026-06-08",
        "marca_comercial": "RoActemra", "principio_activo": "Tocilizumab",
        "tratamiento_id": "TRAT-AR-005", "linea_id": "L-AR-005",
        "visita_id": "VIS-AR-005-02",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-01-10",
        "dosis": "162 mg", "via": "SC",
        "pauta_codigo": "SC-SEM", "pauta_label": "Subcutánea semanal",
        "adherencia_morisky": "Alta", "haq": "1.0", "eva_dolor": "3",
        "hay_efecto_adverso": "TRUE",
        "ea_id": "EA-AR-005-01",
        "ea_descripcion": "Elevación de transaminasas 2x LSN",
        "ea_gravedad": "moderado",
        "farmaco_sospechoso_id": "TRAT-AR-005",
        "farmaco_sospechoso_nombre": "RoActemra (Tocilizumab)",
        "causalidad_naranjo": "probable",
        "causalidad_karch": "probable",
        "accion_ea": "Monitorización mensual de transaminasas",
        "observaciones_generales": "EA hepático bajo seguimiento, continuar RoActemra",
    }))

    # 9. Cambio de pauta
    rows.append(make_row("02_REUMA", 6, "seguimiento", {
        "patologia": "Artritis reumatoide",
        "fecha_acto": "2026-05-30",
        "marca_comercial": "Orencia", "principio_activo": "Abatacept",
        "tratamiento_id": "TRAT-AR-001", "linea_id": "L-AR-001",
        "visita_id": "VIS-AR-001-04",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2025-03-01",
        "dosis": "125 mg", "via": "SC",
        "pauta_codigo": "SC-2SEM", "pauta_label": "Subcutánea cada 2 semanas",
        "pauta_otro_texto": "Ajuste por mejoría mantenida: semanal → bimensual",
        "adherencia_morisky": "Alta", "haq": "0.5", "eva_dolor": "1",
        "respuesta_clinica": "Remisión clínica mantenida, se espacia pauta",
        "incidencias": "Pauta espaciada de semanal a cada 2 semanas por remisión",
    }))

    # 10. Primera visita tras validación
    rows.append(make_row("02_REUMA", 7, "primera_visita", {
        "patologia": "Artritis psoriásica",
        "fecha_acto": "2026-06-05",
        "marca_comercial": "Cosentyx", "principio_activo": "Secukinumab",
        "tipo_validacion": "inicial", "resultado_validacion": "validado",
        "tratamiento_id": "TRAT-AP-001", "linea_id": "L-AP-001",
        "visita_id": "VIS-AP-001-01", "validacion_id": "VAL-AP-001",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-25",
        "dosis": "150 mg", "via": "SC",
        "pauta_codigo": "SC-4SEM", "pauta_label": "Subcutánea cada 4 semanas",
        "adherencia_morisky": "Alta", "haq": "1.8", "eva_dolor": "5", "dlqi": "15",
        "respuesta_clinica": "Primera visita tras inicio, reducción de dolor articular",
    }))

    # 11. Seguimiento con adherencia baja
    rows.append(make_row("02_REUMA", 8, "seguimiento", {
        "patologia": "Artritis reumatoide",
        "fecha_acto": "2026-06-08",
        "marca_comercial": "Olumiant", "principio_activo": "Baricitinib",
        "tratamiento_id": "TRAT-AR-006", "linea_id": "L-AR-006",
        "visita_id": "VIS-AR-006-02",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-03-15",
        "dosis": "4 mg", "via": "oral",
        "pauta_otro_texto": "1 comprimido al día",
        "adherencia_morisky": "Baja", "haq": "2.1", "eva_dolor": "7", "dlqi": "18",
        "respuesta_clinica": "Control articular insuficiente por baja adherencia",
        "incidencias": "Refiere olvidos frecuentes de dosis",
        "observaciones_generales": "Adherencia baja a Olumiant, valorar intervención educativa",
    }))

    # 12. Mismo paciente con varios fármacos activos (Orencia + RoActemra)
    # Patient 0 (FH-SYN-REU-001) already has Orencia. Add RoActemra as concomitante activo.
    rows.append(make_row("02_REUMA", 0, "seguimiento", {
        "patologia": "Artritis reumatoide",
        "fecha_acto": "2026-06-14",
        "marca_comercial": "RoActemra", "principio_activo": "Tocilizumab",
        "tratamiento_id": "TRAT-AR-001-CONC", "linea_id": "L-AR-001-CONC",
        "visita_id": "VIS-AR-001-06",
        "tipo_relacion": "concomitante", "estado_linea": "activo", "es_principal": "FALSE",
        "fecha_inicio": "2026-04-01",
        "dosis": "162 mg", "via": "SC",
        "pauta_codigo": "SC-SEM", "pauta_label": "Subcutánea semanal",
        "adherencia_morisky": "Alta",
        "respuesta_clinica": "Paciente con Orencia (Abatacept) como principal + RoActemra (Tocilizumab) concomitante. Ambos activos.",
        "observaciones_generales": "Varios fármacos activos: Orencia (principal) + RoActemra (concomitante)",
    }))

    return rows


def build_digestivo_rows():
    """10+ rows for Digestivo."""
    rows = []

    # 1. EII pendiente de inducción
    rows.append(make_row("03_DIGESTIVO", 0, "validacion_inicial", {
        "patologia": "Enfermedad de Crohn",
        "fecha_acto": "2026-05-30",
        "marca_comercial": "Stelara", "principio_activo": "Ustekinumab",
        "tipo_validacion": "inicial", "resultado_validacion": "pendiente",
        "requiere_prebiologico": "TRUE",
        "tb_estado": "pendiente", "serologias_estado": "completo", "vacunas_estado": "pendiente",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-30",
        "dosis": "260 mg", "via": "IV",
        "pauta_codigo": "IV-8SEM", "pauta_label": "Intravenosa cada 8 semanas",
        "bloqueantes_validacion": "TB pendiente + vacunas pendientes",
        "observaciones_validacion": "Pendiente de resultado Mantoux y vacunación",
    }))

    # 2. EII en mantenimiento
    rows.append(make_row("03_DIGESTIVO", 1, "seguimiento", {
        "patologia": "Colitis ulcerosa",
        "fecha_acto": "2026-06-10",
        "marca_comercial": "Entyvio", "principio_activo": "Vedolizumab",
        "tratamiento_id": "TRAT-CU-001", "linea_id": "L-CU-001",
        "visita_id": "VIS-CU-001-04",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2025-08-01",
        "dosis": "300 mg", "via": "IV",
        "pauta_codigo": "IV-8SEM", "pauta_label": "Intravenosa cada 8 semanas",
        "adherencia_morisky": "Alta", "haq": "0.5", "eva_dolor": "1", "dlqi": "2",
        "respuesta_clinica": "Remisión clínica mantenida",
    }))

    # 3. Cambio de pauta
    rows.append(make_row("03_DIGESTIVO", 2, "seguimiento", {
        "patologia": "Enfermedad de Crohn",
        "fecha_acto": "2026-06-05",
        "marca_comercial": "Remicade", "principio_activo": "Infliximab",
        "tratamiento_id": "TRAT-EC-001", "linea_id": "L-EC-001",
        "visita_id": "VIS-EC-001-03",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2025-11-01",
        "dosis": "5 mg/kg", "via": "IV",
        "pauta_codigo": "IV-8SEM", "pauta_label": "Intravenosa cada 8 semanas",
        "pauta_otro_texto": "Intensificado de 8 a 6 semanas por pérdida de respuesta",
        "adherencia_morisky": "Alta", "haq": "1.0", "eva_dolor": "3", "dlqi": "8",
        "respuesta_clinica": "Mejoría tras intensificación",
        "incidencias": "Intensificación de pauta de 8 a 6 semanas",
    }))

    # 4. Falta de adherencia
    rows.append(make_row("03_DIGESTIVO", 3, "seguimiento", {
        "patologia": "Colitis ulcerosa",
        "fecha_acto": "2026-06-08",
        "marca_comercial": "Humira", "principio_activo": "Adalimumab",
        "tratamiento_id": "TRAT-CU-002", "linea_id": "L-CU-002",
        "visita_id": "VIS-CU-002-02",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-01-15",
        "dosis": "40 mg", "via": "SC",
        "pauta_codigo": "SC-2SEM", "pauta_label": "Subcutánea cada 2 semanas",
        "adherencia_morisky": "Baja", "haq": "2.0", "eva_dolor": "6", "dlqi": "20",
        "respuesta_clinica": "Brote leve por mala adherencia",
        "incidencias": "Refiere haberse saltado 3 dosis en el último mes",
        "observaciones_generales": "Falta de adherencia a Humira. Valorar intervención educativa.",
    }))

    # 5. Nueva validación por cambio
    rows.append(make_row("03_DIGESTIVO", 4, "nueva_validacion_cambio", {
        "patologia": "Enfermedad de Crohn",
        "fecha_acto": "2026-05-20",
        "marca_comercial": "Skyrizi", "principio_activo": "Risankizumab",
        "validacion_id": "VAL-EC-002",
        "tratamiento_id": "TRAT-EC-002", "linea_id": "L-EC-002",
        "tipo_validacion": "cambio", "resultado_validacion": "validado",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "tipo_movimiento": "cambio_terapeutico",
        "fecha_inicio": "2026-05-25",
        "motivo_inicio_cambio_suspension": "Cambio desde Ustekinumab por respuesta parcial",
        "dosis": "600 mg", "via": "IV",
        "pauta_codigo": "IV-4SEM", "pauta_label": "Intravenosa cada 4 semanas",
        "observaciones_validacion": "Cambio a Risankizumab aprobado por comité",
    }))

    # 6. Suspensión por EA
    rows.append(make_row("03_DIGESTIVO", 5, "suspension", {
        "patologia": "Colitis ulcerosa",
        "fecha_acto": "2026-04-10",
        "marca_comercial": "Entyvio", "principio_activo": "Vedolizumab",
        "tratamiento_id": "TRAT-CU-HIST-001", "linea_id": "L-CU-HIST-001",
        "tipo_relacion": "principal", "estado_linea": "suspendido", "es_principal": "FALSE",
        "tipo_movimiento": "suspension",
        "fecha_inicio": "2025-06-01", "fecha_fin": "2026-04-10",
        "motivo_inicio_cambio_suspension": "Reacción infusional severa",
        "dosis": "300 mg", "via": "IV",
        "pauta_codigo": "IV-8SEM", "pauta_label": "Intravenosa cada 8 semanas",
        "hay_efecto_adverso": "TRUE",
        "ea_id": "EA-CU-HIST-001",
        "ea_descripcion": "Reacción infusional severa (urticaria generalizada + disnea)",
        "ea_gravedad": "grave",
        "farmaco_sospechoso_id": "TRAT-CU-HIST-001",
        "farmaco_sospechoso_nombre": "Entyvio (Vedolizumab)",
        "causalidad_naranjo": "definida",
        "causalidad_karch": "definida",
        "accion_ea": "Suspensión definitiva de Vedolizumab",
        "observaciones_generales": "Suspensión por reacción infusional severa a Vedolizumab",
    }))

    # 7. Fármaco concomitante relevante
    rows.append(make_row("03_DIGESTIVO", 6, "seguimiento", {
        "patologia": "Enfermedad de Crohn",
        "fecha_acto": "2026-06-12",
        "marca_comercial": "Stelara", "principio_activo": "Ustekinumab",
        "tratamiento_id": "TRAT-EC-003", "linea_id": "L-EC-003",
        "visita_id": "VIS-EC-003-02",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-03-01",
        "dosis": "90 mg", "via": "SC",
        "pauta_codigo": "SC-8SEM", "pauta_label": "Subcutánea cada 8 semanas",
        "adherencia_morisky": "Alta", "haq": "0.8", "eva_dolor": "2",
        "respuesta_clinica": "Estable en mantenimiento",
        "observaciones_seguimiento": "Continúa con mesalazina oral como concomitante",
    }))

    # 8. Primera visita
    rows.append(make_row("03_DIGESTIVO", 7, "primera_visita", {
        "patologia": "Enfermedad de Crohn",
        "fecha_acto": "2026-06-01",
        "marca_comercial": "Skyrizi", "principio_activo": "Risankizumab",
        "tipo_validacion": "inicial", "resultado_validacion": "validado",
        "tratamiento_id": "TRAT-EC-002", "linea_id": "L-EC-002",
        "visita_id": "VIS-EC-002-01", "validacion_id": "VAL-EC-002",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-25",
        "dosis": "600 mg", "via": "IV",
        "pauta_codigo": "IV-4SEM", "pauta_label": "Intravenosa cada 4 semanas",
        "adherencia_morisky": "Alta", "dlqi": "14",
        "respuesta_clinica": "Primera visita tras inducción, mejoría sintomática",
    }))

    # 9. Seguimiento con PROMs
    rows.append(make_row("03_DIGESTIVO", 8, "seguimiento", {
        "patologia": "Colitis ulcerosa",
        "fecha_acto": "2026-06-12",
        "marca_comercial": "Entyvio", "principio_activo": "Vedolizumab",
        "tratamiento_id": "TRAT-CU-001", "linea_id": "L-CU-001",
        "visita_id": "VIS-CU-001-05",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2025-08-01",
        "dosis": "300 mg", "via": "IV",
        "pauta_codigo": "IV-8SEM", "pauta_label": "Intravenosa cada 8 semanas",
        "adherencia_morisky": "Alta", "haq": "0.3", "eva_dolor": "1", "dlqi": "1",
        "respuesta_clinica": "Remisión mantenida, PROMs óptimos",
    }))

    # 10. Renovación/continuidad
    rows.append(make_row("03_DIGESTIVO", 9, "seguimiento", {
        "patologia": "Enfermedad de Crohn",
        "fecha_acto": "2026-06-14",
        "marca_comercial": "Stelara", "principio_activo": "Ustekinumab",
        "tratamiento_id": "TRAT-EC-003", "linea_id": "L-EC-003",
        "visita_id": "VIS-EC-003-03",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-03-01",
        "dosis": "90 mg", "via": "SC",
        "pauta_codigo": "SC-8SEM", "pauta_label": "Subcutánea cada 8 semanas",
        "adherencia_morisky": "Alta", "haq": "0.6", "eva_dolor": "2", "dlqi": "4",
        "respuesta_clinica": "Remisión, sin incidencias",
        "observaciones_generales": "Renovación de tratamiento. Todo correcto.",
    }))

    return rows


def build_onco_rows():
    """10+ rows for Oncología covering special drug categories."""
    rows = []

    # 1. Medicación extranjera
    rows.append(make_row("04_ONCO", 0, "validacion_inicial", {
        "patologia": "Melanoma metastásico",
        "fecha_acto": "2026-05-20",
        "marca_comercial": "KEYTRUDA-IMP", "principio_activo": "Pembrolizumab (importación)",
        "codigo_nacional": "EXT-001", "numero_registro": "IMP-ESP-2026-001",
        "source_type": "LOCAL",
        "categoria_farmaco": "biológico-inmunoterapia",
        "tipo_validacion": "inicial", "resultado_validacion": "pendiente",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-20",
        "dosis": "200 mg", "via": "IV",
        "pauta_codigo": "IV-4SEM", "pauta_label": "Intravenosa cada 4 semanas",
        "observaciones_validacion": "Medicación extranjera, pendiente de normalización por farmacia",
        "observaciones_generales": "Fármaco extranjero: Pembrolizumab importado vía medicamentos extranjeros",
    }))

    # 2. Uso compasivo
    rows.append(make_row("04_ONCO", 1, "validacion_inicial", {
        "patologia": "Cáncer de pulmón no microcítico (CPNM) EGFR+",
        "fecha_acto": "2026-06-01",
        "marca_comercial": "Compasivo-OSIM-001", "principio_activo": "Osimertinib (uso compasivo)",
        "codigo_nacional": "COMP-001", "numero_registro": "UC-2026-OSIM-001",
        "source_type": "LOCAL",
        "categoria_farmaco": "pequeña molécula",
        "tipo_validacion": "inicial", "resultado_validacion": "validado",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-06-01",
        "dosis": "80 mg", "via": "oral",
        "pauta_otro_texto": "1 comprimido al día",
        "observaciones_validacion": "Aprobado por comité de uso compasivo",
        "observaciones_generales": "Uso compasivo Osimertinib para CPNM EGFR+ sin alternativa",
    }))

    # 3. Ensayo clínico
    rows.append(make_row("04_ONCO", 2, "seguimiento", {
        "patologia": "Cáncer de mama triple negativo",
        "fecha_acto": "2026-06-10",
        "marca_comercial": "EC-TECENTRIQ-001", "principio_activo": "Atezolizumab (ensayo clínico)",
        "codigo_nacional": "EC-001", "numero_registro": "EC-2026-TNBC-001",
        "source_type": "LOCAL",
        "categoria_farmaco": "biológico-inmunoterapia",
        "tratamiento_id": "TRAT-CM-001", "linea_id": "L-CM-001",
        "visita_id": "VIS-CM-001-03",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-03-15",
        "dosis": "840 mg", "via": "IV",
        "pauta_codigo": "IV-2SEM", "pauta_label": "Intravenosa cada 2 semanas",
        "adherencia_morisky": "Alta",
        "respuesta_clinica": "Respuesta parcial en evaluación de ensayo",
        "observaciones_generales": "Ensayo clínico fase III: Atezolizumab en TNBC",
    }))

    # 4. Fármaco fuera de ficha técnica
    rows.append(make_row("04_ONCO", 3, "nueva_validacion_adicion", {
        "patologia": "Cáncer colorrectal metastásico KRAS wild-type",
        "fecha_acto": "2026-06-05",
        "marca_comercial": "Avastin", "principio_activo": "Bevacizumab",
        "codigo_nacional": "FT-001",
        "source_type": "CIMA",
        "categoria_farmaco": "biológico-antiangiogénico",
        "validacion_id": "VAL-CCR-001",
        "tratamiento_id": "TRAT-CCR-001", "linea_id": "L-CCR-001",
        "tipo_validacion": "adicion", "resultado_validacion": "validado",
        "tipo_relacion": "adicional", "estado_linea": "anadido", "es_principal": "FALSE",
        "tipo_movimiento": "tratamiento_anadido",
        "fecha_inicio": "2026-06-10",
        "motivo_inicio_cambio_suspension": "Adición de Bevacizumab a quimioterapia base",
        "dosis": "5 mg/kg", "via": "IV",
        "pauta_codigo": "IV-2SEM", "pauta_label": "Intravenosa cada 2 semanas",
        "observaciones_validacion": "Bevacizumab fuera de ficha técnica para 3ª línea",
        "observaciones_generales": "Uso fuera de ficha técnica: Bevacizumab en 3ª línea CCR",
    }))

    # 5. Preparación especial
    rows.append(make_row("04_ONCO", 4, "validacion_inicial", {
        "patologia": "Linfoma B difuso de célula grande",
        "fecha_acto": "2026-06-02",
        "marca_comercial": "PREP-ESP-RTX-001", "principio_activo": "Rituximab (preparación especial)",
        "codigo_nacional": "PREP-001",
        "source_type": "LOCAL",
        "categoria_farmaco": "biológico",
        "tipo_validacion": "inicial", "resultado_validacion": "validado",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-06-02",
        "dosis": "375 mg/m2", "via": "IV",
        "pauta_codigo": "IV-4SEM", "pauta_label": "Intravenosa cada 4 semanas",
        "observaciones_validacion": "Preparación especial en cabina de citostáticos",
        "observaciones_generales": "Preparación especial: Rituximab IV con protocolo de premedicación",
    }))

    # 6. Fármaco pendiente de normalización
    rows.append(make_row("04_ONCO", 5, "seguimiento", {
        "patologia": "Cáncer gástrico HER2+",
        "fecha_acto": "2026-06-12",
        "marca_comercial": "PEND-NORM-TRAST-001", "principio_activo": "Trastuzumab (pendiente normalización)",
        "codigo_nacional": "PEND-001",
        "source_type": "LOCAL",
        "categoria_farmaco": "biológico",
        "tratamiento_id": "TRAT-CG-001", "linea_id": "L-CG-001",
        "visita_id": "VIS-CG-001-02",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-05-15",
        "dosis": "6 mg/kg", "via": "IV",
        "pauta_codigo": "IV-4SEM", "pauta_label": "Intravenosa cada 4 semanas",
        "adherencia_morisky": "Alta",
        "respuesta_clinica": "Respuesta parcial",
        "observaciones_generales": "Pendiente de normalización en catálogo de farmacia",
    }))

    # 7. EA moderado
    rows.append(make_row("04_ONCO", 6, "seguimiento", {
        "patologia": "Melanoma metastásico",
        "fecha_acto": "2026-06-08",
        "marca_comercial": "Opdivo", "principio_activo": "Nivolumab",
        "tratamiento_id": "TRAT-MEL-001", "linea_id": "L-MEL-001",
        "visita_id": "VIS-MEL-001-04",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2026-01-10",
        "dosis": "240 mg", "via": "IV",
        "pauta_codigo": "IV-2SEM", "pauta_label": "Intravenosa cada 2 semanas",
        "adherencia_morisky": "Alta",
        "hay_efecto_adverso": "TRUE",
        "ea_id": "EA-MEL-001-01",
        "ea_descripcion": "Colitis inmunomediada grado 2",
        "ea_gravedad": "moderado",
        "farmaco_sospechoso_id": "TRAT-MEL-001",
        "farmaco_sospechoso_nombre": "Opdivo (Nivolumab)",
        "causalidad_naranjo": "probable",
        "causalidad_karch": "probable",
        "accion_ea": "Corticoides orales + suspensión temporal (1 dosis)",
        "observaciones_generales": "Colitis inmunomediada grado 2 por Nivolumab. Resuelto con corticoides.",
    }))

    # 8. Renovación/continuidad
    rows.append(make_row("04_ONCO", 7, "seguimiento", {
        "patologia": "Cáncer de pulmón no microcítico",
        "fecha_acto": "2026-06-14",
        "marca_comercial": "Keytruda", "principio_activo": "Pembrolizumab",
        "tratamiento_id": "TRAT-CPN-001", "linea_id": "L-CPN-001",
        "visita_id": "VIS-CPN-001-06",
        "tipo_relacion": "principal", "estado_linea": "activo", "es_principal": "TRUE",
        "fecha_inicio": "2025-12-01",
        "dosis": "200 mg", "via": "IV",
        "pauta_codigo": "IV-4SEM", "pauta_label": "Intravenosa cada 4 semanas",
        "adherencia_morisky": "Alta", "eva_dolor": "2",
        "respuesta_clinica": "Enfermedad estable mantenida",
        "observaciones_generales": "Renovación de tratamiento. Respuesta mantenida.",
    }))

    # 9. Nueva validación por adición
    rows.append(make_row("04_ONCO", 8, "nueva_validacion_adicion", {
        "patologia": "Cáncer de ovario BRCA+",
        "fecha_acto": "2026-06-05",
        "marca_comercial": "Avastin", "principio_activo": "Bevacizumab",
        "source_type": "CIMA",
        "validacion_id": "VAL-OV-001",
        "tratamiento_id": "TRAT-OV-001", "linea_id": "L-OV-001",
        "tipo_validacion": "adicion", "resultado_validacion": "validado",
        "tipo_relacion": "adicional", "estado_linea": "anadido", "es_principal": "FALSE",
        "tipo_movimiento": "tratamiento_anadido",
        "fecha_inicio": "2026-06-10",
        "motivo_inicio_cambio_suspension": "Adición de antiangiogénico por progresión",
        "dosis": "15 mg/kg", "via": "IV",
        "pauta_codigo": "IV-4SEM", "pauta_label": "Intravenosa cada 4 semanas",
        "observaciones_validacion": "Adición aprobada por comité de tumores",
    }))

    # 10. Suspensión
    rows.append(make_row("04_ONCO", 9, "suspension", {
        "patologia": "Cáncer colorrectal metastásico",
        "fecha_acto": "2026-05-30",
        "marca_comercial": "Avastin", "principio_activo": "Bevacizumab",
        "tratamiento_id": "TRAT-CCR-HIST-001", "linea_id": "L-CCR-HIST-001",
        "tipo_relacion": "principal", "estado_linea": "suspendido", "es_principal": "FALSE",
        "tipo_movimiento": "suspension",
        "fecha_inicio": "2025-07-01", "fecha_fin": "2026-05-30",
        "motivo_inicio_cambio_suspension": "Progresión de enfermedad",
        "dosis": "5 mg/kg", "via": "IV",
        "pauta_codigo": "IV-2SEM", "pauta_label": "Intravenosa cada 2 semanas",
        "observaciones_generales": "Suspensión por progresión oncológica",
    }))

    # 11. Extra: fármaco concomitante en Onco
    rows.append(make_row("04_ONCO", 10, "seguimiento", {
        "patologia": "Cáncer de mama HER2+",
        "fecha_acto": "2026-06-12",
        "marca_comercial": "Herceptin", "principio_activo": "Trastuzumab",
        "tratamiento_id": "TRAT-CM2-001", "linea_id": "L-CM2-001",
        "visita_id": "VIS-CM2-001-02",
        "tipo_relacion": "concomitante", "estado_linea": "activo", "es_principal": "FALSE",
        "fecha_inicio": "2026-04-01",
        "dosis": "6 mg/kg", "via": "IV",
        "pauta_codigo": "IV-4SEM", "pauta_label": "Intravenosa cada 4 semanas",
        "adherencia_morisky": "Alta",
        "respuesta_clinica": "Bien tolerado, sin incidencias",
        "observaciones_generales": "Trastuzumab como concomitante a quimioterapia",
    }))

    return rows


# ── Special Drugs Catalog (05_CATALOGOS) ──────────────────────────
SPECIAL_DRUGS = [
    ("KEYTRUDA-IMP", "Pembrolizumab (importación)", "medicacion_extranjera",
     "Melanoma metastásico. Importación desde UE.", "Importado del Reino Unido por desabastecimiento local.", "2026-01-15", True),
    ("Compasivo-OSIM-001", "Osimertinib (uso compasivo)", "uso_compasivo",
     "CPNM EGFR+ sin alternativa terapéutica.", "Aprobado por comité autonómico uso compasivo.", "2026-02-10", True),
    ("EC-TECENTRIQ-001", "Atezolizumab (ensayo clínico)", "ensayo_clinico",
     "Cáncer mama triple negativo. Estudio IMpassion-131.", "Ensayo fase III promovido por Roche.", "2026-01-20", True),
    ("Avastin-off-label", "Bevacizumab (3ª línea CCR)", "fuera_de_ficha_tecnica",
     "CCR metastásico KRAS wt, 3ª línea.", "Uso off-label aprobado por comité de tumores.", "2026-03-01", True),
    ("PREP-ESP-RTX-001", "Rituximab (preparación especial)", "preparacion_especial",
     "Linfoma B. Preparación en cabina de citostáticos.", "Requiere premeditación y bomba de infusión.", "2026-04-01", True),
    ("PEND-NORM-TRAST-001", "Trastuzumab (pendiente normalización)", "pendiente_normalizacion",
     "Cáncer gástrico HER2+. En proceso de inclusión en catálogo.", "Pendiente de código nacional definitivo.", "2026-05-01", True),
    ("Bevacizumab-OVARIO", "Bevacizumab (cáncer ovario)", "fuera_de_ficha_tecnica",
     "Cáncer de ovario BRCA+ en combinación.", "Aprobado por comité de tumores para esta indicación.", "2026-04-15", True),
    ("Nivolumab-COLITIS", "Nivolumab (colitis inmunomediada)", "otro",
     "Manejo de EA: colitis por inmunoterapia.", "Riesgo de colitis recurrente, monitorización estrecha.", "2026-03-15", True),
    ("Rituximab-LINFOMA-PREP", "Rituximab (preparación especial linfoma)", "preparacion_especial",
     "Protocolo R-CHOP modificado.", "Preparación en cabina con cadena fría.", "2026-02-01", True),
    ("Pembrolizumab-MEL-EXT", "Pembrolizumab (importación melanoma)", "medicacion_extranjera",
     "Melanoma metastásico. Importación temporal.", "Lote importado de Francia por cupo.", "2026-05-10", True),
    ("Trastuzumab-PEND-CAT", "Trastuzumab (pendiente catálogo)", "pendiente_normalizacion",
     "Cáncer gástrico HER2+. Nuevo formato.", "Pendiente de asignación de código nacional.", "2026-05-20", True),
    ("Osimertinib-COMP", "Osimertinib (uso compasivo CPNM)", "uso_compasivo",
     "Segunda línea CPNM EGFR T790M.", "Renovación uso compasivo aprobada 2026.", "2026-01-05", True),
]


def create_synthetic_excel():
    """Create the synthetic Excel workbook."""
    # Load base template for styles and catalogs
    wb = load_workbook(TEMPLATE_PATH)

    for sheet_name, svc in SERVICES.items():
        ws = wb[sheet_name]
        # Clear existing data rows (keep header row 1)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # Build data rows
        if sheet_name == "01_DERMA":
            data = build_derma_rows()
        elif sheet_name == "02_REUMA":
            data = build_reuma_rows()
        elif sheet_name == "03_DIGESTIVO":
            data = build_digestivo_rows()
        elif sheet_name == "04_ONCO":
            data = build_onco_rows()
        else:
            data = []

        for i, row in enumerate(data):
            for j, val in enumerate(row):
                cell = ws.cell(row=i + 2, column=j + 1, value=val)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(vertical="center")
                cell.border = BORDER_THIN

    # Populate 05_CATALOGOS special drugs
    ws_cat = wb["05_CATALOGOS"]
    # Find the special drugs section by looking for headers
    # The catalog sheet has labeled sections; append to special drugs area
    # Look for an existing row count and append after
    cat_data_start = ws_cat.max_row + 2  # Leave a blank row

    # Write section header
    section_cell = ws_cat.cell(row=cat_data_start, column=1,
                               value="FÁRMACOS ESPECIALES (sintéticos)")
    section_cell.font = Font(name="Calibri", bold=True, size=11, color="2F5496")

    # Headers
    spec_headers = ["marca_nombre_visible", "principio_activo", "categoria_especial",
                    "indicacion_uso", "observaciones", "fecha_alta_catalogo", "activo"]
    for j, h in enumerate(spec_headers):
        cell = ws_cat.cell(row=cat_data_start + 1, column=j + 1, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN

    for i, drug in enumerate(SPECIAL_DRUGS):
        for j, val in enumerate(drug):
            cell = ws_cat.cell(row=cat_data_start + 2 + i, column=j + 1, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = BORDER_THIN

    wb.save(OUTPUT_PATH)
    print(f"Synthetic Excel created: {OUTPUT_PATH}")
    print(f"  Services: {len(SERVICES)} sheets")
    print(f"  DERMA rows: {len(build_derma_rows())}")
    print(f"  REUMA rows: {len(build_reuma_rows())}")
    print(f"  DIGESTIVO rows: {len(build_digestivo_rows())}")
    print(f"  ONCO rows: {len(build_onco_rows())}")
    print(f"  Special drugs: {len(SPECIAL_DRUGS)}")
    return True


if __name__ == "__main__":
    create_synthetic_excel()
