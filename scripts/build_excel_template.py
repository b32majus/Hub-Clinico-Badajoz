#!/usr/bin/env python3
"""Generate WO8.1a — Plantilla Excel operativa FH."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

# ── Columnas de servicio (bloques A-H) ──────────────────────────
SERVICE_COLS = [
    # A. Identificación paciente
    ("patient_id", "A", "string", True),
    ("cip_demo_o_hash", "A", "string", True),
    ("nhc_o_codigo_interno", "A", "string", False),
    ("fecha_nacimiento_o_edad", "A", "date/number", True),
    ("sexo", "A", "string", True),
    ("servicio_origen", "A", "string", True),
    ("patologia_indicacion", "A", "string", True),
    # B. Acto farmacéutico
    ("fecha_acto", "B", "date", True),
    ("tipo_acto_fh", "B", "string", True),
    ("visita_id", "B", "string", False),
    ("validacion_id", "B", "string", False),
    ("tratamiento_id", "B", "string", False),
    ("linea_id", "B", "string", False),
    ("profesional_fh", "B", "string", True),
    ("estado_registro", "B", "string", True),
    # C. Medicamento / línea terapéutica
    ("marca_comercial", "C", "string", True),
    ("principio_activo", "C", "string", True),
    ("codigo_nacional", "C", "string", False),
    ("numero_registro", "C", "string", False),
    ("source_type", "C", "string", False),
    ("categoria_farmaco", "C", "string", False),
    ("tipo_relacion", "C", "string", True),
    ("estado_linea", "C", "string", True),
    ("tipo_movimiento", "C", "string", False),
    ("es_principal", "C", "boolean", True),
    ("fecha_inicio", "C", "date", True),
    ("fecha_fin", "C", "date", False),
    ("motivo_inicio_cambio_suspension", "C", "string", False),
    # D. Pauta y administración
    ("dosis_presentacion", "D", "string", True),
    ("via", "D", "string", True),
    ("pauta_codigo", "D", "string", False),
    ("pauta_label", "D", "string", False),
    ("pauta_otro_texto", "D", "string", False),
    # E. Validación farmacoterapéutica
    ("tipo_validacion", "E", "string", False),
    ("resultado_validacion", "E", "string", False),
    ("requiere_prebiologico", "E", "boolean", False),
    ("tb_estado", "E", "string", False),
    ("serologias_estado", "E", "string", False),
    ("vacunas_estado", "E", "string", False),
    ("bloqueantes_validacion", "E", "string", False),
    ("observaciones_validacion", "E", "string", False),
    # F. Seguimiento
    ("adherencia_morisky", "F", "string", False),
    ("haq", "F", "number", False),
    ("eva_dolor", "F", "number", False),
    ("dlqi", "F", "number", False),
    ("respuesta_clinica", "F", "string", False),
    ("incidencias", "F", "string", False),
    ("observaciones_seguimiento", "F", "string", False),
    # G. Seguridad / EA
    ("hay_efecto_adverso", "G", "boolean", True),
    ("ea_id", "G", "string", False),
    ("ea_descripcion", "G", "string", False),
    ("ea_gravedad", "G", "string", False),
    ("farmaco_sospechoso_id", "G", "string", False),
    ("farmaco_sospechoso_nombre", "G", "string", False),
    ("causalidad_naranjo", "G", "string", False),
    ("causalidad_karch", "G", "string", False),
    ("accion_ea", "G", "string", False),
    # H. Trazabilidad
    ("created_at", "H", "datetime", True),
    ("updated_at", "H", "datetime", False),
    ("demo_flag", "H", "boolean", True),
    ("observaciones_generales", "H", "string", False),
]

HEADER_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SERVICES = {
    "01_DERMA": "Dermatología",
    "02_REUMA": "Reumatología",
    "03_DIGESTIVO": "Digestivo",
    "04_ONCO": "Oncología",
}

# ── Estilos ─────────────────────────────────────────────────────
BLOCK_FILLS = {
    "A": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "B": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "C": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "D": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "E": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "F": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "G": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    "H": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}

BLOCK_LABELS = {
    "A": "A. Identificación paciente",
    "B": "B. Acto farmacéutico",
    "C": "C. Medicamento / línea terapéutica",
    "D": "D. Pauta y administración",
    "E": "E. Validación farmacoterapéutica",
    "F": "F. Seguimiento",
    "G": "G. Seguridad / EA",
    "H": "H. Trazabilidad",
}

# Columna P0 crítica
P0_COLS = {"patient_id", "cip_demo_o_hash", "servicio_origen", "fecha_acto",
           "tipo_acto_fh", "profesional_fh", "marca_comercial", "principio_activo",
           "tipo_relacion", "estado_linea", "es_principal", "fecha_inicio",
           "dosis_presentacion", "via", "hay_efecto_adverso", "demo_flag", "created_at"}

# ── Listas controladas ──────────────────────────────────────────
LISTS = {
    "tipo_acto_fh": [
        "validacion_inicial", "primera_visita", "seguimiento",
        "nueva_validacion_cambio", "nueva_validacion_adicion",
        "suspension", "cambio_pauta", "efecto_adverso",
        "renovacion_continuidad", "otro",
    ],
    "estado_registro": ["activo", "completado", "pendiente_revision"],
    "sexo": ["Hombre", "Mujer", "No binario", "No especificado"],
    "source_type": ["CIMA", "LOCAL", "DEMO", "EXCEL", "MANUAL"],
    "categoria_farmaco": ["Biológico", "Biosimilar", "Pequeña molécula", "Otro"],
    "tipo_relacion": ["principal", "adicional", "concomitante", "historico", "exposicion"],
    "estado_linea": ["activo", "suspendido", "historico", "finalizado", "anadido"],
    "tipo_movimiento": ["sin_cambios", "cambio_terapeutico", "tratamiento_anadido", "suspension", "cambio_pauta"],
    "via": ["SC", "IV", "VO", "IM", "Tópica", "Intraarticular", "Otra"],
    "tipo_validacion": ["inicial", "cambio", "adicion", "renovacion"],
    "resultado_validacion": ["validado", "pendiente", "rechazado", "no_aplica"],
    "requiere_prebiologico": ["TRUE", "FALSE"],
    "tb_estado": ["Negativo", "Positivo", "Pendiente", "No realizado", "No aplica"],
    "serologias_estado": ["Negativo", "Positivo", "Pendiente", "No realizado", "No aplica"],
    "vacunas_estado": ["Completo", "Incompleto", "Pendiente", "No aplica"],
    "adherencia_morisky": ["Alta", "Media", "Baja", "No evaluada"],
    "hay_efecto_adverso": ["TRUE", "FALSE"],
    "ea_gravedad": ["Grave", "Moderado", "Leve"],
    "causalidad_naranjo": ["Definitiva", "Probable", "Posible", "Dudosa", "No evaluada"],
    "causalidad_karch": ["Definitiva", "Probable", "Posible", "Dudosa", "No evaluada"],
    "demo_flag": ["TRUE", "FALSE"],
}

ESPECIAL_CATEGORIES = [
    "fuera_de_ficha_tecnica", "ensayo_clinico", "uso_compasivo",
    "medicacion_extranjera", "preparacion_especial",
    "pendiente_normalizacion", "otro",
]


def build_service_sheet(ws, service_name, service_label):
    """Create a service sheet with headers and formatting."""
    headers = [col[0] for col in SERVICE_COLS]
    blocks = [col[1] for col in SERVICE_COLS]
    required = [col[3] for col in SERVICE_COLS]

    # Write headers
    for c, (hdr, blk, _, _) in enumerate(SERVICE_COLS, 1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.font = HEADER_FONT
        fill = BLOCK_FILLS.get(blk, HEADER_FILL)
        cell.fill = fill
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER_THIN

    # Column widths
    col_widths = {
        1: 14, 2: 16, 3: 18, 4: 12, 5: 10, 6: 14, 7: 20,  # A
        8: 12, 9: 22, 10: 14, 11: 14, 12: 16, 13: 14, 14: 20, 15: 16,  # B
        16: 18, 17: 18, 18: 14, 19: 14, 20: 12, 21: 16, 22: 14, 23: 14,
        24: 18, 25: 10, 26: 12, 27: 12, 28: 30,  # C
        29: 22, 30: 8, 31: 16, 32: 16, 33: 22,  # D
        34: 16, 35: 16, 36: 14, 37: 12, 38: 12, 39: 12, 40: 18, 41: 24,  # E
        42: 16, 43: 8, 44: 8, 45: 8, 46: 20, 47: 22, 48: 24,  # F
        49: 14, 50: 14, 51: 24, 52: 12, 53: 18, 54: 22, 55: 16, 56: 14, 57: 22,  # G
        58: 18, 59: 18, 60: 10, 61: 28,  # H
    }
    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SERVICE_COLS))}1"

    # Add a few demo rows
    if service_label == "Dermatología":
        demo_rows = [
            ["CIP-DEMO-FH-001", "CIP-DEMO-FH-001", "", "45", "Mujer",
             "Dermatología", "Psoriasis",
             "2026-06-01", "seguimiento", "VIS-001", "", "TRAT-FH-001-A", "BIO-FH-001-L1",
             "Ana Farmacéutica", "completado",
             "Humira", "Adalimumab", "", "", "CIMA", "Biológico",
             "principal", "activo", "sin_cambios", "TRUE",
             "2026-01-15", "", "Respuesta adecuada. Continuar.",
             "40 mg", "SC", "CADA_2_SEMANAS", "Cada 2 semanas", "",
             "", "", "", "", "", "", "", "",
             "Alta", "", "", "", "", "", "",
             "FALSE", "", "", "", "", "", "", "", "",
             "2026-06-01T10:00:00", "", "TRUE", "Fila demo"],
        ]
        for r_idx, row in enumerate(demo_rows, 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = Font(size=9, italic=True, color="666666")
                cell.border = BORDER_THIN

    return True


# ═════════════════════════════════════════════════════════════════
# CREATE SHEETS
# ═════════════════════════════════════════════════════════════════

# Remove default sheet
default_ws = wb.active

# ── Service sheets ─────────────────────────────────────────────
for sheet_name, service_label in SERVICES.items():
    ws = wb.create_sheet(title=sheet_name)
    build_service_sheet(ws, sheet_name, service_label)
    # Pre-fill servicio_origen column (F) with the service label
    for row in range(2, 5):
        ws.cell(row=row, column=6, value=service_label)

# Remove the default sheet created by Workbook
wb.remove(default_ws)

# ── 05_CATALOGOS ───────────────────────────────────────────────
cat_ws = wb.create_sheet(title="05_CATALOGOS")

# Section 1: Listas desplegables
cat_ws.cell(row=1, column=1, value="LISTAS DESPLEGABLES").font = Font(bold=True, size=12, color="2F5496")
row = 3
for list_name, values in LISTS.items():
    cat_ws.cell(row=row, column=1, value=list_name).font = Font(bold=True, size=10)
    cat_ws.cell(row=row, column=1).fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for v_idx, val in enumerate(values):
        cat_ws.cell(row=row + 1 + v_idx, column=1, value=val)
    row += len(values) + 2

# Section 2: Fármacos especiales
row = max(row + 2, row)
cat_ws.cell(row=row, column=1, value="CATÁLOGO DE FÁRMACOS ESPECIALES").font = Font(bold=True, size=12, color="2F5496")
row += 1
esp_headers = ["marca_nombre_visible", "principio_activo", "categoria_especial",
               "indicacion_uso", "observaciones", "fecha_alta_catalogo", "activo"]
for c, hdr in enumerate(esp_headers, 1):
    cell = cat_ws.cell(row=row, column=c, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER_THIN
row += 1
# Example rows
esp_examples = [
    ["Fármaco en investigación X", "Principio activo X", "ensayo_clinico",
     "Artritis Reumatoide refractaria", "Ensayo fase III - hospital", "2026-06-01", "TRUE"],
    ["Medicamento extranjero Y", "", "medicacion_extranjera",
     "LES", "Importado vía medicamentos extranjeros", "2026-05-15", "TRUE"],
]
for eg in esp_examples:
    for c, val in enumerate(eg, 1):
        cell = cat_ws.cell(row=row, column=c, value=val)
        cell.font = Font(size=9, italic=True, color="666666")
        cell.border = BORDER_THIN
    row += 1

# Categories list
row += 1
cat_ws.cell(row=row, column=1, value="Categorías de fármaco especial:").font = Font(bold=True, size=10)
row += 1
for cat in ESPECIAL_CATEGORIES:
    cat_ws.cell(row=row, column=1, value=cat)
    row += 1

cat_ws.column_dimensions["A"].width = 35
cat_ws.column_dimensions["B"].width = 25
cat_ws.column_dimensions["C"].width = 30
cat_ws.column_dimensions["D"].width = 35
cat_ws.column_dimensions["E"].width = 35
cat_ws.column_dimensions["F"].width = 18
cat_ws.column_dimensions["G"].width = 10

# ── 99_CONFIG_EXPORT_MAP ───────────────────────────────────────
map_ws = wb.create_sheet(title="99_CONFIG_EXPORT_MAP")
map_headers = ["columna_excel", "bloque", "entidad_destino", "campo_destino",
               "obligatorio", "tipo_dato", "comentario"]
for c, hdr in enumerate(map_headers, 1):
    cell = map_ws.cell(row=1, column=c, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER_THIN

# Map P0 columns to entities
ENTITY_MAP = {
    "pacientes": {
        "patient_id": ("A", "pacientes", "patient_id", "Sí", "string", "PK"),
        "cip_demo_o_hash": ("A", "pacientes", "cip_demo_o_hash", "Sí", "string", ""),
        "servicio_origen": ("A", "pacientes", "servicio_origen", "Sí", "string", "Partición de hoja"),
        "patologia_indicacion": ("A", "pacientes", "patologia", "Sí", "string", ""),
        "sexo": ("A", "pacientes", "sexo", "Sí", "string", ""),
    },
    "lineas_tratamiento": {
        "marca_comercial": ("C", "lineas_tratamiento", "marca_comercial", "Sí", "string", "Nombre principal"),
        "principio_activo": ("C", "lineas_tratamiento", "principio_activo", "Sí", "string", "Campo secundario"),
        "tipo_relacion": ("C", "lineas_tratamiento", "tipo_relacion", "Sí", "string", ""),
        "estado_linea": ("C", "lineas_tratamiento", "estado_linea", "Sí", "string", ""),
        "es_principal": ("C", "lineas_tratamiento", "es_principal", "Sí", "boolean", ""),
        "fecha_inicio": ("C", "lineas_tratamiento", "fecha_inicio", "Sí", "date", ""),
        "fecha_fin": ("C", "lineas_tratamiento", "fecha_fin", "No", "date", "Vacío si activa"),
        "linea_id": ("B", "lineas_tratamiento", "linea_id", "No", "string", "ID lógico"),
        "tratamiento_id": ("B", "lineas_tratamiento", "tratamiento_id", "No", "string", "PK tratamiento"),
    },
    "visitas_seguimiento": {
        "fecha_acto": ("B", "visitas", "fecha_visita", "Sí", "date", "Fecha del acto"),
        "tipo_acto_fh": ("B", "visitas", "tipo_acto_fh", "Sí", "string", "Clasifica el acto"),
        "profesional_fh": ("B", "visitas", "profesional_fh", "Sí", "string", ""),
    },
    "efectos_adversos": {
        "hay_efecto_adverso": ("G", "ea", "hay_ea", "Sí", "boolean", "Flag de presencia"),
        "ea_descripcion": ("G", "ea", "descripcion", "No", "string", ""),
        "ea_gravedad": ("G", "ea", "gravedad", "No", "string", ""),
        "farmaco_sospechoso_nombre": ("G", "ea", "farmaco_sospechoso", "No", "string", ""),
        "causalidad_naranjo": ("G", "ea", "causalidad_naranjo", "No", "string", ""),
    },
    "general": {
        "demo_flag": ("H", "general", "demo_flag", "Sí", "boolean", "TRUE = demo"),
        "created_at": ("H", "general", "created_at", "Sí", "datetime", ""),
    },
}

r = 2
for entity, fields in ENTITY_MAP.items():
    for col_name, (blk, dest_entity, dest_field, req, dtype, comment) in fields.items():
        map_ws.cell(row=r, column=1, value=col_name).border = BORDER_THIN
        map_ws.cell(row=r, column=2, value=blk).border = BORDER_THIN
        map_ws.cell(row=r, column=3, value=dest_entity).border = BORDER_THIN
        map_ws.cell(row=r, column=4, value=dest_field).border = BORDER_THIN
        map_ws.cell(row=r, column=5, value=req).border = BORDER_THIN
        map_ws.cell(row=r, column=6, value=dtype).border = BORDER_THIN
        map_ws.cell(row=r, column=7, value=comment).border = BORDER_THIN
        r += 1

map_ws.column_dimensions["A"].width = 30
map_ws.column_dimensions["B"].width = 10
map_ws.column_dimensions["C"].width = 22
map_ws.column_dimensions["D"].width = 22
map_ws.column_dimensions["E"].width = 12
map_ws.column_dimensions["F"].width = 12
map_ws.column_dimensions["G"].width = 25
map_ws.freeze_panes = "A2"

# ── Save ────────────────────────────────────────────────────────
output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "templates")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "farmacia_excel_operativo_FH_WO8_v1.xlsx")
wb.save(output_path)
print(f"✅ Plantilla creada: {output_path}")
print(f"   Hojas: {wb.sheetnames}")
