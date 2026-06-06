#!/usr/bin/env python3
"""
Genera la base demo v2 canónica para Reuma:
- 321 columnas históricas del maestro (AR) intactas.
- 176 columnas v2 (322-497) según docs/ORDEN_COLUMNAS_EXCEL_REUMA_V2.md.
- 5 hojas clínicas: AR, ESPA, APS, LES, SJOGREN (497 columnas cada una).
- Cohorte demo cliente: 50 pacientes / 200 visitas (10 pacientes por patología, 4 visitas por paciente).
- Hojas auxiliares copiadas del maestro: Profesionales y Fármacos.
"""

from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Sequence, Tuple

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:
    raise SystemExit("ERROR: openpyxl no está instalado. Ejecuta: pip install openpyxl") from exc


PROJECT_DIR = Path(__file__).resolve().parents[1]
MASTER_XLSX = PROJECT_DIR / "Hub_Clinico_Maestro.xlsx"
OUTPUT_XLSX = PROJECT_DIR / "data" / "Hub_Clinico_Maestro_V2_DEMO.xlsx"
ORDER_MD = PROJECT_DIR / "docs" / "ORDEN_COLUMNAS_EXCEL_REUMA_V2.md"
REPORT_MD = PROJECT_DIR / "docs" / "REPORTE_DIFERENCIAS_EXCEL_DEMO_V2.md"

CLINICAL_SHEETS = ["AR", "ESPA", "APS", "LES", "SJOGREN"]
MASTER_BASE_SHEETS = ["AR", "ESPA", "APS"]
FINAL_COLUMN_COUNT = 497
HISTORICAL_COLUMN_COUNT = 321
V2_START = 322
V2_END = 497
PATIENTS_PER_PATHOLOGY = 10
VISITS_PER_PATIENT = 4

DAPSA_HEADERS = [
    "DAPSA_Result",
    "DAPSA_NAD68",
    "DAPSA_NAT66",
    "DAPSA_EVA_Dolor_Paciente",
    "DAPSA_EVA_Global_Paciente",
    "DAPSA_PCR",
]

PREBIO_STATES = ["NO_EVALUADO", "EN_CURSO", "APTO", "NO_APTO"]


@dataclass
class ValidationCheck:
    label: str
    ok: bool
    detail: str = ""


def normalize_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_header(ws) -> List[str]:
    return [normalize_header(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]


def parse_v2_headers_from_contract(path: Path) -> List[str]:
    if not path.exists():
        raise FileNotFoundError(f"No existe contrato de columnas: {path}")

    line_re = re.compile(r"^\|\s*(\d+)\s*\|\s*`([^`]+)`\s*\|")
    positions: Dict[int, str] = {}

    for raw_line in path.read_text(encoding="utf-8").splitlines():
        m = line_re.match(raw_line.strip())
        if not m:
            continue
        pos = int(m.group(1))
        col = m.group(2).strip()
        if V2_START <= pos <= V2_END:
            positions[pos] = col

    expected_positions = list(range(V2_START, V2_END + 1))
    missing = [p for p in expected_positions if p not in positions]
    if missing:
        raise ValueError(f"Contrato incompleto: faltan posiciones v2 {missing[:5]} ... total {len(missing)}")

    v2_headers = [positions[p] for p in expected_positions]
    expected_v2_count = V2_END - V2_START + 1
    if len(v2_headers) != expected_v2_count:
        raise ValueError(f"Se esperaban {expected_v2_count} columnas v2 y se obtuvieron {len(v2_headers)}")

    return v2_headers


def load_master_historical_headers(master_path: Path) -> Tuple[List[str], Dict[str, List[str]]]:
    if not master_path.exists():
        raise FileNotFoundError(f"No existe Excel maestro: {master_path}")

    wb = load_workbook(master_path, read_only=True, data_only=True)
    headers_by_sheet: Dict[str, List[str]] = {}
    for sheet in MASTER_BASE_SHEETS:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Falta hoja {sheet} en maestro.")
        headers_by_sheet[sheet] = read_header(wb[sheet])
    return headers_by_sheet["AR"], headers_by_sheet


def build_empty_row(headers: Sequence[str]) -> Dict[str, str]:
    return {h: "" for h in headers}


def set_if_exists(row: Dict[str, str], value, *keys: str) -> None:
    for key in keys:
        if key in row:
            row[key] = value


def parse_demo_number(value):
    if value is None or value == "":
        return None
    raw = str(value).strip().replace(",", ".")
    try:
        return float(raw)
    except ValueError:
        return None


def format_demo_number(value, decimals: int = 1) -> str:
    if value is None:
        return ""
    text = f"{value:.{decimals}f}"
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def normalize_eva_0_10(value):
    number = parse_demo_number(value)
    if number is None:
        return None
    if number > 10:
        return number / 10
    return number


def first_number(*values):
    for value in values:
        number = parse_demo_number(value)
        if number is not None:
            return number
    return None


def finalize_dapsa_contract(row: Dict[str, str]) -> None:
    if not any(header in row for header in DAPSA_HEADERS):
        return

    pathology = str(row.get("Diagnostico_Primario") or row.get("Diagnostico_Principal") or "").strip().lower()
    if pathology != "aps":
        for header in DAPSA_HEADERS:
            if header in row:
                row[header] = "NA"
        return

    nad68 = first_number(row.get("DAPSA_NAD68"), row.get("NAD_Total"))
    nat66 = first_number(row.get("DAPSA_NAT66"), row.get("NAT_Total"))
    eva_dolor = normalize_eva_0_10(row.get("DAPSA_EVA_Dolor_Paciente") or row.get("EVA_Dolor"))
    eva_global = normalize_eva_0_10(row.get("DAPSA_EVA_Global_Paciente") or row.get("EVA_Global"))
    pcr_mgl = first_number(row.get("DAPSA_PCR"), row.get("PCR"))

    values = [nad68, nat66, eva_dolor, eva_global, pcr_mgl]
    if any(value is None for value in values):
        return

    if "DAPSA_NAD68" in row:
        row["DAPSA_NAD68"] = format_demo_number(nad68, 0)
    if "DAPSA_NAT66" in row:
        row["DAPSA_NAT66"] = format_demo_number(nat66, 0)
    if "DAPSA_EVA_Dolor_Paciente" in row:
        row["DAPSA_EVA_Dolor_Paciente"] = format_demo_number(eva_dolor, 1)
    if "DAPSA_EVA_Global_Paciente" in row:
        row["DAPSA_EVA_Global_Paciente"] = format_demo_number(eva_global, 1)
    if "DAPSA_PCR" in row:
        row["DAPSA_PCR"] = format_demo_number(pcr_mgl, 1)
    if "DAPSA_Result" in row:
        dapsa = nad68 + nat66 + eva_dolor + eva_global + (pcr_mgl / 10)
        row["DAPSA_Result"] = format_demo_number(dapsa, 1)


def base_visit_row(headers: Sequence[str], patient_id: str, name: str, sex: str, pathology: str) -> Dict[str, str]:
    row = build_empty_row(headers)
    set_if_exists(row, patient_id, "ID_Paciente")
    set_if_exists(row, name, "Nombre_Paciente")
    set_if_exists(row, sex, "Sexo")
    set_if_exists(row, pathology, "Diagnostico_Primario", "Diagnostico_Principal")
    set_if_exists(row, "NO", "Comorbilidad_DM", "Comorbilidad_ECV")
    set_if_exists(row, "ND", "Toxico_Tabaco", "Toxico_Alcohol")
    return row


def apply_payload_to_row(row: Dict[str, str], payload: Dict[str, str]) -> None:
    for key, value in payload.items():
        if key in row:
            row[key] = value

    if "Decision_Terapeutica" in payload and "Decision_Terapeutica" not in row:
        tipo = str(payload.get("Tipo_Visita", "")).strip().lower()
        decision = payload.get("Decision_Terapeutica", "")
        if tipo == "primera" and "Decision_Terapeutica_PV" in row:
            row["Decision_Terapeutica_PV"] = decision
        if tipo == "seguimiento" and "Decision_Terapeutica_SEG" in row:
            row["Decision_Terapeutica_SEG"] = decision

    finalize_dapsa_contract(row)


def normalize_id(value) -> str:
    return str(value).strip() if value is not None else ""


def select_master_patients(master_path: Path, sheet: str, count: int) -> List[Tuple[str, str, str]]:
    wb = load_workbook(master_path, read_only=True, data_only=True)
    ws = wb[sheet]
    headers = read_header(ws)
    idx = {h: i for i, h in enumerate(headers)}

    id_idx = idx.get("ID_Paciente")
    name_idx = idx.get("Nombre_Paciente")
    sex_idx = idx.get("Sexo")
    if id_idx is None:
        raise ValueError(f"Hoja {sheet}: falta columna ID_Paciente")

    by_id: Dict[str, Tuple[str, str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = normalize_id(row[id_idx] if id_idx < len(row) else "")
        if not pid:
            continue
        if pid in by_id:
            continue
        name = normalize_id(row[name_idx] if name_idx is not None and name_idx < len(row) else "")
        sex = normalize_id(row[sex_idx] if sex_idx is not None and sex_idx < len(row) else "")
        if not name:
            name = f"Paciente {pid}"
        if sex not in {"F", "M"}:
            sex = "F" if len(by_id) % 2 == 0 else "M"
        by_id[pid] = (pid, name, sex)

    selected = sorted(by_id.values(), key=lambda t: t[0])[:count]
    if len(selected) < count:
        raise ValueError(f"Hoja {sheet}: no hay suficientes pacientes únicos ({len(selected)}/{count}).")
    return selected


def build_demo_patient_catalog(master_path: Path) -> Dict[str, List[Tuple[str, str, str]]]:
    catalog = {
        "AR": select_master_patients(master_path, "AR", PATIENTS_PER_PATHOLOGY),
        "ESPA": select_master_patients(master_path, "ESPA", PATIENTS_PER_PATHOLOGY),
        "APS": select_master_patients(master_path, "APS", PATIENTS_PER_PATHOLOGY),
        "LES": [],
        "SJOGREN": [],
    }

    for i in range(1, PATIENTS_PER_PATHOLOGY + 1):
        les_id = f"LES-2026-{i:03d}"
        sj_id = f"SJOGREN-2026-{i:03d}"
        les_sex = "F" if i % 4 else "M"
        sj_sex = "F" if i % 5 else "M"
        catalog["LES"].append((les_id, f"Paciente {les_id}", les_sex))
        catalog["SJOGREN"].append((sj_id, f"Paciente {sj_id}", sj_sex))

    return catalog


def score_bucket(base: float, visit_idx: int, slope: float, floor: float) -> float:
    return round(max(floor, base - (visit_idx * slope)), 1)


def make_visit_payload(
    pathology: str,
    patient_idx: int,
    visit_idx: int,
    visit_date: date,
    diagnosis_date: date,
) -> Dict[str, str]:
    payload: Dict[str, str] = {}
    payload["Fecha_Visita"] = visit_date.isoformat()
    payload["Tipo_Visita"] = "primera" if visit_idx == 0 else "seguimiento"
    payload["Fecha_Diagnostico"] = diagnosis_date.isoformat()
    payload["Fecha_Proxima_Revision"] = (visit_date + timedelta(days=120)).isoformat()
    payload["Peso"] = str(58 + ((patient_idx * 3) % 24))
    payload["Talla"] = str(155 + ((patient_idx * 2) % 20))
    payload["TA"] = f"{112 + (patient_idx % 6) * 4}/{68 + (patient_idx % 5) * 2}"
    payload["Comorbilidad_HTA"] = "SI" if patient_idx % 3 == 0 else "NO"
    payload["Comorbilidad_DLP"] = "SI" if patient_idx % 4 == 0 else "NO"
    payload["Comorbilidad_Obesidad"] = "SI" if patient_idx % 5 == 0 else "NO"

    final_state = PREBIO_STATES[patient_idx % len(PREBIO_STATES)]
    if visit_idx < 2 and final_state in {"APTO", "NO_APTO"}:
        status = "EN_CURSO"
    elif visit_idx == 0 and final_state == "NO_EVALUADO":
        status = "NO_EVALUADO"
    else:
        status = final_state
    payload["Estado_Prebiologico_Final"] = status
    if status in {"APTO", "NO_APTO"}:
        payload["Fecha_Validacion_Prebiologico"] = (visit_date - timedelta(days=3)).isoformat()
        payload["Vacunacion_OK"] = "SI"
    else:
        payload["Vacunacion_OK"] = "NO" if status == "NO_APTO" else "SI"
    payload["Vacunacion_Revisada"] = "SI"

    if visit_idx == 0:
        payload["Decision_Terapeutica"] = "iniciar"
    elif visit_idx == 1:
        payload["Decision_Terapeutica"] = "cambiar"
    else:
        payload["Decision_Terapeutica"] = "continuar"

    if pathology == "AR":
        pcr = score_bucket(18 - patient_idx * 0.6, visit_idx, 3.2, 1.5)
        vsg = score_bucket(40 - patient_idx * 0.8, visit_idx, 6.5, 8.0)
        das28 = score_bucket(6.4 - patient_idx * 0.1, visit_idx, 1.1, 2.0)
        cdai = score_bucket(31 - patient_idx * 0.7, visit_idx, 6.0, 3.0)
        sdai = round(cdai + (pcr / 10), 1)
        rapid3 = score_bucket(17 - patient_idx * 0.4, visit_idx, 3.5, 2.0)
        haq = score_bucket(2.2 - patient_idx * 0.06, visit_idx, 0.35, 0.2)
        payload.update(
            {
                "PCR": format_demo_number(pcr, 1),
                "VSG": format_demo_number(vsg, 1),
                "NAD28": format_demo_number(max(1, 10 - visit_idx * 2 - patient_idx * 0.2), 0),
                "NAT28": format_demo_number(max(1, 8 - visit_idx * 2 - patient_idx * 0.2), 0),
                "EVA_Dolor": format_demo_number(max(20, 82 - visit_idx * 16 - patient_idx * 2), 0),
                "EVA_Global": format_demo_number(max(18, 78 - visit_idx * 15 - patient_idx * 2), 0),
                "EVA_Medico": format_demo_number(max(18, 75 - visit_idx * 14 - patient_idx * 2), 0),
                "DAS28_CRP_Result": format_demo_number(das28, 1),
                "CDAI_Result": format_demo_number(cdai, 1),
                "SDAI_Result": format_demo_number(sdai, 1),
                "RAPID3_Score": format_demo_number(rapid3, 1),
                "HAQ_Total": format_demo_number(haq, 1),
            }
        )
        payload["Tratamiento_Actual"] = [
            "Metotrexato 20 mg/sem",
            "Metotrexato 20 mg/sem + Adalimumab 40 mg/14d",
            "Baricitinib 4 mg/día",
            "Baricitinib 4 mg/día",
        ][visit_idx]
    elif pathology == "ESPA":
        pcr = score_bucket(14 - patient_idx * 0.4, visit_idx, 2.6, 1.2)
        vsg = score_bucket(30 - patient_idx * 0.5, visit_idx, 4.5, 7.0)
        basdai = score_bucket(7.4 - patient_idx * 0.2, visit_idx, 1.5, 1.6)
        asdas = score_bucket(3.9 - patient_idx * 0.08, visit_idx, 0.75, 1.2)
        haq = score_bucket(1.6 - patient_idx * 0.05, visit_idx, 0.2, 0.2)
        payload.update(
            {
                "PCR": format_demo_number(pcr, 1),
                "VSG": format_demo_number(vsg, 1),
                "BASDAI_Result": format_demo_number(basdai, 1),
                "ASDAS_CRP_Result": format_demo_number(asdas, 1),
                "HAQ_Total": format_demo_number(haq, 1),
                "BASFI_Result": format_demo_number(score_bucket(6.8 - patient_idx * 0.2, visit_idx, 1.1, 1.5), 1),
            }
        )
        payload["Tratamiento_Actual"] = [
            "Naproxeno 500 mg/12h",
            "Secukinumab 150 mg/mes",
            "Secukinumab 150 mg/mes",
            "Secukinumab 150 mg/mes",
        ][visit_idx]
    elif pathology == "APS":
        nad68 = score_bucket(16 - patient_idx * 0.4, visit_idx, 3.2, 1.0)
        nat66 = score_bucket(12 - patient_idx * 0.35, visit_idx, 2.4, 1.0)
        eva_dolor = score_bucket(8.2 - patient_idx * 0.15, visit_idx, 1.3, 1.6)
        eva_global = score_bucket(7.8 - patient_idx * 0.12, visit_idx, 1.2, 1.5)
        pcr = score_bucket(16 - patient_idx * 0.5, visit_idx, 3.0, 1.5)
        payload.update(
            {
                "PCR": format_demo_number(pcr, 1),
                "VSG": format_demo_number(score_bucket(28 - patient_idx * 0.4, visit_idx, 4.8, 8.0), 1),
                "NAD_Total": format_demo_number(nad68, 0),
                "NAT_Total": format_demo_number(nat66, 0),
                "EVA_Dolor": format_demo_number(eva_dolor, 1),
                "EVA_Global": format_demo_number(eva_global, 1),
                "PASI_Score": format_demo_number(score_bucket(18 - patient_idx * 0.4, visit_idx, 4.3, 1.0), 1),
                "BSA_Percentage": format_demo_number(score_bucket(24 - patient_idx * 0.6, visit_idx, 5.2, 2.0), 1),
                "LEI_Score": format_demo_number(score_bucket(7 - patient_idx * 0.2, visit_idx, 1.6, 0.0), 0),
                "HAQ_Total": format_demo_number(score_bucket(1.8 - patient_idx * 0.04, visit_idx, 0.3, 0.2), 1),
                "RAPID3_Score": format_demo_number(score_bucket(15 - patient_idx * 0.35, visit_idx, 3.0, 2.0), 1),
                "DAPSA_NAD68": format_demo_number(nad68, 0),
                "DAPSA_NAT66": format_demo_number(nat66, 0),
                "DAPSA_EVA_Dolor_Paciente": format_demo_number(eva_dolor, 1),
                "DAPSA_EVA_Global_Paciente": format_demo_number(eva_global, 1),
                "DAPSA_PCR": format_demo_number(pcr, 1),
            }
        )
        payload["Tratamiento_Actual"] = [
            "Metotrexato 20 mg/sem",
            "Metotrexato 20 mg/sem + Secukinumab 150 mg/mes",
            "Secukinumab 150 mg/mes",
            "Secukinumab 150 mg/mes",
        ][visit_idx]
    elif pathology == "LES":
        pcr = score_bucket(12 - patient_idx * 0.3, visit_idx, 2.5, 1.0)
        vsg = score_bucket(36 - patient_idx * 0.6, visit_idx, 5.5, 7.0)
        sledai = score_bucket(14 - patient_idx * 0.5, visit_idx, 3.6, 1.0)
        slicc = score_bucket(1 + patient_idx * 0.25, visit_idx, -0.2, 0.0)
        pred = score_bucket(25 - patient_idx * 0.4, visit_idx, 6.5, 2.5)
        payload.update(
            {
                "PCR": format_demo_number(pcr, 1),
                "VSG": format_demo_number(vsg, 1),
                "SLEDAI": format_demo_number(sledai, 0),
                "SLEDAI_2K": format_demo_number(sledai, 0),
                "SLICC_SDI": format_demo_number(slicc, 0),
                "Dosis_Prednisona": format_demo_number(pred, 1),
                "PCR_LES": format_demo_number(pcr, 1),
                "VSG_LES": format_demo_number(vsg, 1),
            }
        )
        payload["Tratamiento_Actual"] = [
            "Hidroxicloroquina 400 mg/día + Prednisona",
            "Hidroxicloroquina 400 mg/día + Micofenolato 2 g/día",
            "Hidroxicloroquina 400 mg/día + Micofenolato 2 g/día",
            "Hidroxicloroquina 400 mg/día + Micofenolato 1 g/día",
        ][visit_idx]
    else:
        pcr = score_bucket(9 - patient_idx * 0.25, visit_idx, 1.8, 1.0)
        vsg = score_bucket(22 - patient_idx * 0.4, visit_idx, 3.2, 6.0)
        essdai = score_bucket(16 - patient_idx * 0.45, visit_idx, 3.5, 2.0)
        esspri_seq = score_bucket(7.5 - patient_idx * 0.12, visit_idx, 1.1, 1.5)
        esspri_fat = score_bucket(7.0 - patient_idx * 0.10, visit_idx, 1.0, 1.4)
        esspri_dol = score_bucket(6.8 - patient_idx * 0.10, visit_idx, 0.9, 1.2)
        esspri = round((esspri_seq + esspri_fat + esspri_dol) / 3, 1)
        payload.update(
            {
                "PCR": format_demo_number(pcr, 1),
                "VSG": format_demo_number(vsg, 1),
                "ESSDAI_Result": format_demo_number(essdai, 0),
                "ESSPRI_Sequedad": format_demo_number(esspri_seq, 1),
                "ESSPRI_Fatiga": format_demo_number(esspri_fat, 1),
                "ESSPRI_Dolor": format_demo_number(esspri_dol, 1),
                "ESSPRI_Result": format_demo_number(esspri, 1),
                "EVA_Sequedad_Oral": format_demo_number(score_bucket(8.0 - patient_idx * 0.12, visit_idx, 1.2, 1.0), 1),
                "EVA_Sequedad_Ocular": format_demo_number(score_bucket(7.8 - patient_idx * 0.11, visit_idx, 1.1, 1.0), 1),
                "EVA_Fatiga_Sjogren": format_demo_number(esspri_fat, 1),
                "EVA_Dolor_Sjogren": format_demo_number(esspri_dol, 1),
                "EVA_Global_Sjogren": format_demo_number(score_bucket(7.4 - patient_idx * 0.10, visit_idx, 1.0, 1.0), 1),
                "PCR_Sjogren": format_demo_number(pcr, 1),
                "VSG_Sjogren": format_demo_number(vsg, 1),
            }
        )
        payload["Tratamiento_Actual"] = [
            "Pilocarpina 5 mg/8h",
            "Pilocarpina 5 mg/8h + Hidroxicloroquina 400 mg/día",
            "Pilocarpina 5 mg/8h + Hidroxicloroquina 400 mg/día",
            "Pilocarpina 5 mg/8h + Rituximab 1 g/6 meses",
        ][visit_idx]

    payload["Fecha_Inicio_Tratamiento"] = (
        visit_date.isoformat() if visit_idx <= 1 else (visit_date - timedelta(days=150)).isoformat()
    )
    if visit_idx == 1:
        payload["Cambio_Motivo"] = "Ajuste por actividad clínica persistente"
    if visit_idx == 2 and patient_idx % 3 == 0:
        payload["Cambio_Efectos_Adversos"] = "Intolerancia digestiva"
    if visit_idx == 3 and patient_idx % 4 == 0:
        payload["Comentarios_Adicionales"] = "Mejoría sostenida en seguimiento."

    return payload


def build_clinical_rows(headers: List[str]) -> Dict[str, List[Dict[str, str]]]:
    rows: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    catalog = build_demo_patient_catalog(MASTER_XLSX)

    pathology_code = {
        "AR": "ar",
        "ESPA": "espa",
        "APS": "aps",
        "LES": "les",
        "SJOGREN": "sjogren",
    }
    pathology_offset = {"AR": 0, "ESPA": 20, "APS": 40, "LES": 60, "SJOGREN": 80}

    for sheet in CLINICAL_SHEETS:
        patients = catalog[sheet]
        for pidx, (pid, name, sex) in enumerate(patients):
            first_visit = date(2025, 1, 15) + timedelta(days=pathology_offset[sheet] + (pidx * 9))
            diagnosis_date = first_visit - timedelta(days=200 + (pidx * 7))
            visit_dates = [first_visit + timedelta(days=i * 150) for i in range(VISITS_PER_PATIENT)]
            for vidx, visit_date in enumerate(visit_dates):
                row = base_visit_row(headers, pid, name, sex, pathology_code[sheet])
                payload = make_visit_payload(sheet, pidx, vidx, visit_date, diagnosis_date)
                apply_payload_to_row(row, payload)
                rows[sheet].append(row)

    return rows


def copy_sheet_values(source_wb, target_wb, sheet_name: str):
    if sheet_name not in source_wb.sheetnames:
        raise ValueError(f"Falta hoja auxiliar '{sheet_name}' en {MASTER_XLSX.name}")
    src = source_wb[sheet_name]
    dst = target_wb.create_sheet(sheet_name)
    for row in src.iter_rows(values_only=True):
        dst.append(list(row))
    return dst


def build_auxiliary_sheets(wb: Workbook) -> None:
    master_wb = load_workbook(MASTER_XLSX, read_only=True, data_only=True)
    ws_prof = copy_sheet_values(master_wb, wb, "Profesionales")
    copy_sheet_values(master_wb, wb, "Fármacos")

    headers = read_header(ws_prof)
    idx_name = headers.index("Nombre") if "Nombre" in headers else None
    idx_cargo = headers.index("Cargo") if "Cargo" in headers else None
    if idx_name is None or idx_cargo is None:
        raise ValueError("La hoja Profesionales del maestro no contiene columnas esperadas Nombre/Cargo.")

    existing_names = set()
    for row in ws_prof.iter_rows(min_row=2, values_only=True):
        if row is None or idx_name >= len(row):
            continue
        value = row[idx_name]
        if value is not None:
            existing_names.add(str(value).strip().casefold())

    if "raúl veroz".casefold() not in existing_names:
        new_row = ["" for _ in headers]
        new_row[idx_name] = "Raúl Veroz"
        new_row[idx_cargo] = "Reumatólogo"
        ws_prof.append(new_row)


def write_workbook(headers: List[str], rows_by_sheet: Dict[str, List[Dict[str, str]]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in CLINICAL_SHEETS:
        ws = wb.create_sheet(sheet)
        ws.append(headers)
        for row_dict in rows_by_sheet[sheet]:
            ws.append([row_dict.get(h, "") for h in headers])

    build_auxiliary_sheets(wb)
    wb.save(output_path)


def validate_headers_no_empty_no_dup(headers: List[str]) -> Tuple[bool, bool, List[str], List[str]]:
    empties = [f"index {idx + 1}" for idx, h in enumerate(headers) if not h]
    counter = Counter(headers)
    dups = [k for k, v in counter.items() if v > 1 and k]
    return len(empties) == 0, len(dups) == 0, empties, dups


def run_post_generation_validations(
    demo_path: Path,
    master_headers: List[str],
    rows_by_sheet: Dict[str, List[Dict[str, str]]],
) -> Tuple[List[ValidationCheck], Dict[str, Dict[str, int]], Dict[str, int]]:
    wb = load_workbook(demo_path, read_only=True, data_only=True)
    checks: List[ValidationCheck] = []
    sheet_shape: Dict[str, Dict[str, int]] = {}

    for sheet in CLINICAL_SHEETS:
        ws = wb[sheet]
        headers = read_header(ws)
        n_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        n_cols = len(headers)
        sheet_shape[sheet] = {"columns": n_cols, "rows": n_rows}
        checks.append(ValidationCheck(f"{sheet} = {FINAL_COLUMN_COUNT} columnas", n_cols == FINAL_COLUMN_COUNT, f"{n_cols}"))
        checks.append(
            ValidationCheck(
                f"{sheet} = {PATIENTS_PER_PATHOLOGY * VISITS_PER_PATIENT} visitas",
                n_rows == PATIENTS_PER_PATHOLOGY * VISITS_PER_PATIENT,
                f"{n_rows}",
            )
        )

    for sheet in MASTER_BASE_SHEETS:
        headers = read_header(wb[sheet])
        same = headers[:HISTORICAL_COLUMN_COUNT] == master_headers
        checks.append(
            ValidationCheck(
                f"primeras 321 columnas de {sheet} coinciden con maestro",
                same,
                "OK" if same else "Diferencia detectada",
            )
        )

    for sheet in CLINICAL_SHEETS:
        headers = read_header(wb[sheet])
        no_empty, no_dups, empties, dups = validate_headers_no_empty_no_dup(headers)
        checks.append(ValidationCheck(f"{sheet} sin cabeceras vacías", no_empty, "" if no_empty else ", ".join(empties[:5])))
        checks.append(ValidationCheck(f"{sheet} sin cabeceras duplicadas", no_dups, "" if no_dups else ", ".join(dups[:5])))

    aps_dapsa_populated = all(str(r.get("DAPSA_Result", "")).strip() not in ("", "NA", "ND") for r in rows_by_sheet.get("APS", []))
    checks.append(
        ValidationCheck(
            "APS DAPSA poblado",
            aps_dapsa_populated,
            f"{sum(str(r.get('DAPSA_Result', '')).strip() not in ('', 'NA', 'ND') for r in rows_by_sheet.get('APS', []))}/{len(rows_by_sheet.get('APS', []))} visitas",
        )
    )

    non_aps_na = True
    for sheet, rows in rows_by_sheet.items():
        if sheet == "APS":
            continue
        for r in rows:
            if any(str(r.get(header, "")).strip() != "NA" for header in DAPSA_HEADERS):
                non_aps_na = False
                break
        if not non_aps_na:
            break
    checks.append(ValidationCheck("DAPSA = NA en no APs", non_aps_na))

    patient_visits: Dict[str, int] = {}
    patient_types_ok = True
    unique_patients_by_sheet: Dict[str, int] = {}

    for sheet, rows in rows_by_sheet.items():
        by_patient: Dict[str, List[Dict[str, str]]] = defaultdict(list)
        for row in rows:
            by_patient[str(row.get("ID_Paciente", "?"))].append(row)
        unique_patients_by_sheet[sheet] = len(by_patient)
        for pid, p_rows in by_patient.items():
            patient_visits[pid] = len(p_rows)
            if len(p_rows) != VISITS_PER_PATIENT:
                patient_types_ok = False
            first_count = sum(str(r.get("Tipo_Visita", "")).strip().lower() == "primera" for r in p_rows)
            seg_count = sum(str(r.get("Tipo_Visita", "")).strip().lower() == "seguimiento" for r in p_rows)
            if first_count != 1 or seg_count != 3:
                patient_types_ok = False
            dates = [str(r.get("Fecha_Visita", "")) for r in p_rows]
            if dates != sorted(dates):
                patient_types_ok = False

    checks.append(
        ValidationCheck(
            "cada paciente tiene 1 primera + 3 seguimientos",
            patient_types_ok,
            f"{len(patient_visits)} pacientes, {sum(patient_visits.values())} visitas total",
        )
    )
    checks.append(
        ValidationCheck(
            "10 pacientes por patología",
            all(v == PATIENTS_PER_PATHOLOGY for v in unique_patients_by_sheet.values()),
            str(unique_patients_by_sheet),
        )
    )
    checks.append(
        ValidationCheck(
            "cohorte total = 50 pacientes / 200 visitas",
            len(patient_visits) == (PATIENTS_PER_PATHOLOGY * len(CLINICAL_SHEETS))
            and sum(patient_visits.values()) == (PATIENTS_PER_PATHOLOGY * VISITS_PER_PATIENT * len(CLINICAL_SHEETS)),
            f"{len(patient_visits)} / {sum(patient_visits.values())}",
        )
    )

    checks.append(ValidationCheck("sin hoja Prebiologico", "Prebiologico" not in wb.sheetnames))
    checks.append(ValidationCheck("sin hoja Solicitud_FH", "Solicitud_FH" not in wb.sheetnames and "Solicitud FH" not in wb.sheetnames))

    return checks, sheet_shape, patient_visits


def write_diff_report(
    report_path: Path,
    master_path: Path,
    output_path: Path,
    master_headers: List[str],
    v2_headers: List[str],
    checks: List[ValidationCheck],
    sheet_shape: Dict[str, Dict[str, int]],
    patient_visits: Dict[str, int],
) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    total_patients = len(patient_visits)
    total_visits = sum(patient_visits.values())

    lines: List[str] = []
    lines.append("# Reporte Diferencias Excel Demo v2")
    lines.append("")
    lines.append(f"_Generado automáticamente: {now}_")
    lines.append("")
    lines.append("## Fuente canónica")
    lines.append("1. Excel maestro original (`Hub_Clinico_Maestro.xlsx`) para columnas 1-321.")
    lines.append(f"2. `docs/ORDEN_COLUMNAS_EXCEL_REUMA_V2.md` para columnas {V2_START}-{V2_END}.")
    lines.append("3. `modules/exportManager.js` como verificación secundaria del orden v2.")
    lines.append("")
    lines.append("## Archivos")
    lines.append(f"- Maestro leído: `{master_path}`")
    lines.append(f"- Demo generado: `{output_path}`")
    lines.append("")
    lines.append("## Estructura final")
    lines.append(f"- Columnas históricas intactas: `{len(master_headers)}`")
    lines.append(f"- Columnas v2 añadidas: `{len(v2_headers)}`")
    lines.append(f"- Total por hoja clínica: `{len(master_headers) + len(v2_headers)}`")
    lines.append("")
    lines.append("## Cohorte demo cliente")
    lines.append(f"- Pacientes totales: `{total_patients}`")
    lines.append(f"- Visitas totales: `{total_visits}`")
    lines.append(f"- Regla longitudinal: `10 pacientes por patología` y `4 visitas por paciente (1 primera + 3 seguimientos)`")
    lines.append("- Nota: estos conteos describen este dataset demo y no representan un límite funcional de producción.")
    lines.append("")
    lines.append("## AUDIT-FIX-2 ejecutado — DAPSA incorporado al contrato APs")
    lines.append("- Contrato: 497 columnas por hoja clínica.")
    lines.append("- DAPSA (492-497) poblado en APs y `NA` en no APs.")
    lines.append("")
    lines.append("## Resumen por hoja")
    lines.append("")
    lines.append("| Hoja | Columnas | Filas |")
    lines.append("|---|---:|---:|")
    for sheet in CLINICAL_SHEETS:
        shape = sheet_shape.get(sheet, {"columns": 0, "rows": 0})
        lines.append(f"| {sheet} | {shape['columns']} | {shape['rows']} |")
    lines.append("| Profesionales | (auxiliar) | (copiada del maestro + Raúl Veroz) |")
    lines.append("| Fármacos | (auxiliar) | (copiada del maestro) |")
    lines.append("")
    lines.append("## Pacientes demo longitudinales")
    lines.append("")
    for pid, n in sorted(patient_visits.items()):
        lines.append(f"- `{pid}`: {n} visitas")
    lines.append("")
    lines.append("## Validaciones automáticas")
    lines.append("")
    for check in checks:
        mark = "[x]" if check.ok else "[ ]"
        detail = f" — {check.detail}" if check.detail else ""
        lines.append(f"- {mark} {check.label}{detail}")
    lines.append("")
    lines.append("## Diferencias conocidas")
    lines.append("- LES/SJOGREN no existen en el maestro original; se construyen con las 321 históricas + 176 v2.")
    lines.append("- No se crea hoja `Prebiologico` obligatoria en esta demo (prebiológico embebido por visita).")
    lines.append("- No se crea ninguna columna/hoja de `Solicitud FH` (salida derivada TXT).")
    lines.append("")
    lines.append(f"## Columnas v2 añadidas ({V2_START}-{V2_END})")
    lines.append("")
    lines.append(f"Total: {len(v2_headers)}")
    lines.append("")
    lines.append("```text")
    for idx, header in enumerate(v2_headers, start=V2_START):
        lines.append(f"{idx:03d} {header}")
    lines.append("```")
    lines.append("")

    report_path.write_text("\n".join(lines), encoding="utf-8")


def print_console_summary(
    output_path: Path,
    checks: List[ValidationCheck],
    sheet_shape: Dict[str, Dict[str, int]],
    patient_visits: Dict[str, int],
) -> None:
    print(f"[OK] Demo generado: {output_path}")
    print("")
    print("Resumen de hojas clínicas:")
    for sheet in CLINICAL_SHEETS:
        shape = sheet_shape[sheet]
        print(f"  - {sheet}: {shape['rows']} filas, {shape['columns']} columnas")
    print("")
    print(f"Pacientes demo totales: {len(patient_visits)}")
    print(f"Visitas demo totales: {sum(patient_visits.values())}")
    print("")
    print("Validaciones:")
    for check in checks:
        state = "OK " if check.ok else "FAIL"
        detail = f" ({check.detail})" if check.detail else ""
        print(f"  [{state}] {check.label}{detail}")


def main() -> None:
    ar_headers, headers_by_sheet = load_master_historical_headers(MASTER_XLSX)

    for sheet, headers in headers_by_sheet.items():
        if len(headers) != HISTORICAL_COLUMN_COUNT:
            raise ValueError(f"{sheet} no tiene {HISTORICAL_COLUMN_COUNT} columnas (tiene {len(headers)}).")
    if headers_by_sheet["ESPA"] != ar_headers or headers_by_sheet["APS"] != ar_headers:
        raise ValueError("AR/ESPA/APS no tienen cabeceras históricas idénticas en el maestro.")

    v2_headers = parse_v2_headers_from_contract(ORDER_MD)
    clinical_headers = ar_headers + v2_headers
    if len(clinical_headers) != FINAL_COLUMN_COUNT:
        raise ValueError(f"Cabecera final inválida: {len(clinical_headers)} columnas (esperado {FINAL_COLUMN_COUNT}).")

    rows_by_sheet = build_clinical_rows(clinical_headers)
    write_workbook(clinical_headers, rows_by_sheet, OUTPUT_XLSX)

    checks, sheet_shape, patient_visits = run_post_generation_validations(
        OUTPUT_XLSX,
        ar_headers,
        rows_by_sheet,
    )
    write_diff_report(
        REPORT_MD,
        MASTER_XLSX,
        OUTPUT_XLSX,
        ar_headers,
        v2_headers,
        checks,
        sheet_shape,
        patient_visits,
    )
    print_console_summary(OUTPUT_XLSX, checks, sheet_shape, patient_visits)


if __name__ == "__main__":
    main()
